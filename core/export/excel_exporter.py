"""Workbook-writing helpers for the recap Excel template."""

from __future__ import annotations

import os
import sys
from copy import copy
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from job_cost_tool.core.config import ConfigLoader




def _export_debug_enabled() -> bool:
    """Return True when export debug instrumentation is enabled."""
    return str(os.getenv("JOB_COST_TOOL_EXPORT_DEBUG") or "").strip().casefold() not in {"", "0", "false", "no", "off"}


def _debug_log(message: str) -> None:
    """Emit guarded export debug logging."""
    if _export_debug_enabled():
        print(f"[export-debug][excel_exporter] {message}", file=sys.stderr, flush=True)

_SECTION_LABELS = {
    "materials": "Material",
    "subcontractors": "Subcontractor",
    "permits & fees": "Permits & Fees",
    "police detail": "Police Detail",
}


def export_to_excel(template_path: str, output_path: str, recap_payload: dict[str, Any]) -> None:
    """Write recap payload values into a copy of the existing recap template workbook."""
    template = Path(template_path).expanduser().resolve()
    output = Path(output_path).expanduser().resolve()
    loader = ConfigLoader()
    mapping = loader.get_recap_template_map()
    labor_row_slots = loader.get_labor_row_slots()
    equipment_row_slots = loader.get_equipment_row_slots()
    worksheet_name = str(mapping.get("worksheet_name", "")).strip()
    if not worksheet_name:
        raise ValueError("Recap template map is missing worksheet_name.")

    workbook, worksheet = _load_template_workbook(template, worksheet_name)
    _validate_output_path(template, output)
    _validate_section_capacities(recap_payload, mapping)

    _clear_fixed_inputs(worksheet, labor_row_slots, equipment_row_slots)
    _clear_list_section(worksheet, "materials", mapping["materials_section"])
    _clear_list_section(worksheet, "subcontractors", mapping["subcontractors_section"])
    _clear_list_section(worksheet, "permits & fees", mapping["permits_fees_section"])
    _clear_list_section(worksheet, "police detail", mapping["police_detail_section"])
    _clear_header_fields(worksheet, mapping.get("header_fields", {}))

    _write_header_fields(worksheet, recap_payload.get("header", {}), mapping.get("header_fields", {}))
    _write_fixed_row_labels(worksheet, labor_row_slots, "labor")
    _write_fixed_row_labels(worksheet, equipment_row_slots, "equipment")
    _write_labor_values(worksheet, recap_payload.get("labor", {}), labor_row_slots)
    _write_labor_rates(worksheet, recap_payload.get("labor_rates", {}), labor_row_slots)
    _write_equipment_values(worksheet, recap_payload.get("equipment", {}), equipment_row_slots)
    _write_equipment_rates(worksheet, recap_payload.get("equipment_rates", {}), equipment_row_slots)
    _write_list_section(worksheet, recap_payload.get("materials", []), mapping["materials_section"], "materials")
    _write_list_section(
        worksheet,
        recap_payload.get("subcontractors", []),
        mapping["subcontractors_section"],
        "subcontractors",
    )
    _write_list_section(
        worksheet,
        recap_payload.get("permits_fees", []),
        mapping["permits_fees_section"],
        "permits & fees",
    )
    _write_list_section(
        worksheet,
        recap_payload.get("police_detail", []),
        mapping["police_detail_section"],
        "police detail",
    )
    _write_sales_tax_area(
        worksheet,
        mapping.get("sales_tax_area", {}),
        mapping.get("materials_section", {}),
    )
    _write_summary_totals_area(
        worksheet,
        recap_payload.get("project_management_total"),
        mapping.get("sales_tax_area", {}),
        mapping.get("materials_section", {}),
        mapping.get("subcontractors_section", {}),
        mapping.get("permits_fees_section", {}),
        mapping.get("police_detail_section", {}),
    )

    _debug_log(
        f"module={__file__} template={template} output={output} payload_pm_total={recap_payload.get('project_management_total')!r} pre_save_E59={worksheet['E59'].value!r} pre_save_F59={worksheet['F59'].value!r}"
    )

    try:
        workbook.save(output)
    except PermissionError as exc:
        if output.exists():
            raise ValueError(
                "The output file is currently open. Please close it and try again."
            ) from exc
        raise ValueError(f"Permission denied while saving recap workbook to '{output}'.") from exc
    except OSError as exc:
        if output.exists():
            raise ValueError(
                "The output file is currently open. Please close it and try again."
            ) from exc
        raise ValueError(f"Failed to save recap workbook '{output}': {exc}") from exc

    if _export_debug_enabled():
        reopened_workbook = load_workbook(output)
        reopened_sheet = reopened_workbook[worksheet_name]
        _debug_log(
            f"post_save_output={output} post_save_E59={reopened_sheet['E59'].value!r} post_save_F59={reopened_sheet['F59'].value!r}"
        )


def _load_template_workbook(template: Path, worksheet_name: str) -> tuple[Any, Worksheet]:
    """Load and validate the configured recap template workbook."""
    if not template.is_file():
        raise FileNotFoundError(f"Recap template workbook was not found: {template}")

    try:
        workbook = load_workbook(template)
    except (InvalidFileException, BadZipFile) as exc:
        raise ValueError(f"Recap template workbook is not a valid Excel file: {template}") from exc
    except Exception as exc:
        raise ValueError(f"Failed to load recap template workbook '{template}': {exc}") from exc

    if worksheet_name not in workbook.sheetnames:
        raise ValueError(f"Recap template worksheet '{worksheet_name}' was not found in '{template}'.")

    return workbook, workbook[worksheet_name]


def _validate_output_path(template: Path, output: Path) -> None:
    """Validate the chosen output path before saving the recap workbook."""
    if template == output:
        raise ValueError("Output path must be different from the recap template path.")
    if not output.parent.exists():
        raise FileNotFoundError(f"Output folder does not exist: {output.parent}")


def _validate_section_capacities(recap_payload: dict[str, Any], mapping: dict[str, Any]) -> None:
    """Fail early when a variable-length section would overflow the template."""
    section_key_map = {
        "materials": "materials_section",
        "subcontractors": "subcontractors_section",
        "permits_fees": "permits_fees_section",
        "police_detail": "police_detail_section",
    }
    display_name_map = {
        "materials": "Material",
        "subcontractors": "Subcontractor",
        "permits_fees": "Permits & Fees",
        "police_detail": "Police Detail",
    }

    for payload_key, mapping_key in section_key_map.items():
        rows = recap_payload.get(payload_key, []) or []
        section_name = payload_key.replace("_", " ")
        start_row, end_row, _ = _get_section_bounds(section_name, mapping[mapping_key])
        capacity = end_row - start_row + 1
        required = len(rows)
        if required > capacity:
            display_name = display_name_map[payload_key]
            raise ValueError(
                f"{display_name} section exceeds template capacity ({capacity} rows available, {required} required)."
            )


def _clear_fixed_inputs(
    worksheet: Worksheet,
    labor_row_slots: dict[str, Any],
    equipment_row_slots: dict[str, Any],
) -> None:
    """Clear fixed writable input cells before writing a new export."""
    for slot_info in labor_row_slots.values():
        if not isinstance(slot_info, dict):
            continue
        row_mapping = slot_info.get("mapping", {})
        if not isinstance(row_mapping, dict):
            continue
        label_cell = _get_fixed_row_label_cell(slot_info, "labor")
        if label_cell:
            worksheet[str(label_cell)].value = None
        for key in ("st_hours", "ot_hours", "dt_hours", "st_rate", "ot_rate", "dt_rate"):
            cell_ref = row_mapping.get(key)
            if cell_ref:
                worksheet[str(cell_ref)].value = None

    for slot_info in equipment_row_slots.values():
        if not isinstance(slot_info, dict):
            continue
        row_mapping = slot_info.get("mapping", {})
        if not isinstance(row_mapping, dict):
            continue
        label_cell = _get_fixed_row_label_cell(slot_info, "equipment")
        if label_cell:
            worksheet[str(label_cell)].value = None
        for key in ("hours_qty", "rate"):
            cell_ref = row_mapping.get(key)
            if cell_ref:
                worksheet[str(cell_ref)].value = None


def _clear_header_fields(worksheet: Worksheet, header_mapping: dict[str, Any]) -> None:
    """Clear only mapped header input cells before writing known values."""
    for mapping_entry in header_mapping.values():
        if not isinstance(mapping_entry, dict):
            continue
        cell_ref = mapping_entry.get("cell")
        if cell_ref:
            worksheet[str(cell_ref)].value = None


def _clear_list_section(worksheet: Worksheet, section_name: str, section_mapping: dict[str, Any]) -> None:
    """Clear only the writable cells in a bounded variable-length section."""
    start_row, end_row, columns = _get_section_bounds(section_name, section_mapping)
    for row in range(start_row, end_row + 1):
        for column_letter in columns.values():
            worksheet[f"{column_letter}{row}"].value = None


def _write_header_fields(
    worksheet: Worksheet,
    header_payload: dict[str, Any],
    header_mapping: dict[str, Any],
) -> None:
    """Write known header values into their configured input cells."""
    for header_key, value in header_payload.items():
        if value in {None, ""}:
            continue
        mapping_entry = header_mapping.get(header_key)
        if not isinstance(mapping_entry, dict) or "cell" not in mapping_entry:
            continue
        worksheet[str(mapping_entry["cell"])].value = value


def _write_fixed_row_labels(
    worksheet: Worksheet,
    row_slots: dict[str, Any],
    family: str,
) -> None:
    """Write active slot labels into the fixed recap row label cells."""
    for slot_info in row_slots.values():
        if not isinstance(slot_info, dict):
            continue
        label_cell = _get_fixed_row_label_cell(slot_info, family)
        label = str(slot_info.get("label") or "").strip() if bool(slot_info.get("active")) else ""
        worksheet[str(label_cell)].value = label or None


def _write_labor_values(
    worksheet: Worksheet,
    labor_payload: dict[str, dict[str, int | float]],
    labor_row_slots: dict[str, Any],
) -> None:
    """Write aggregated labor hours into fixed recap labor rows keyed by slot id."""
    for slot_id, totals in labor_payload.items():
        mapping_entry = _get_fixed_row_mapping(labor_row_slots, slot_id, "labor")

        for hour_type, mapping_key in (("ST", "st_hours"), ("OT", "ot_hours"), ("DT", "dt_hours")):
            cell_ref = mapping_entry.get(mapping_key)
            if not cell_ref:
                raise ValueError(
                    f"Recap template map for labor slot '{slot_id}' is missing '{mapping_key}'."
                )
            value = totals.get(hour_type, 0)
            worksheet[str(cell_ref)].value = None if _is_zero(value) else value


def _write_labor_rates(
    worksheet: Worksheet,
    labor_rate_payload: dict[str, dict[str, int | float]],
    labor_row_slots: dict[str, Any],
) -> None:
    """Write configured labor rates into recap labor rate input cells keyed by slot id."""
    for slot_id, rate_values in labor_rate_payload.items():
        mapping_entry = _get_fixed_row_mapping(labor_row_slots, slot_id, "labor")

        for payload_key, mapping_key in (("ST", "st_rate"), ("OT", "ot_rate"), ("DT", "dt_rate")):
            if payload_key not in rate_values:
                continue
            cell_ref = mapping_entry.get(mapping_key)
            if not cell_ref:
                raise ValueError(
                    f"Recap template map for labor slot '{slot_id}' is missing '{mapping_key}'."
                )
            worksheet[str(cell_ref)].value = rate_values[payload_key]


def _write_equipment_values(
    worksheet: Worksheet,
    equipment_payload: dict[str, int | float],
    equipment_row_slots: dict[str, Any],
) -> None:
    """Write aggregated equipment hours or quantities into fixed recap rows keyed by slot id."""
    for slot_id, value in equipment_payload.items():
        mapping_entry = _get_fixed_row_mapping(equipment_row_slots, slot_id, "equipment")
        cell_ref = mapping_entry.get("hours_qty")
        if not cell_ref:
            raise ValueError(
                f"Recap template map for equipment slot '{slot_id}' is missing 'hours_qty'."
            )
        worksheet[str(cell_ref)].value = None if _is_zero(value) else value


def _write_equipment_rates(
    worksheet: Worksheet,
    equipment_rate_payload: dict[str, int | float],
    equipment_row_slots: dict[str, Any],
) -> None:
    """Write configured equipment rates into recap equipment rate input cells keyed by slot id."""
    for slot_id, value in equipment_rate_payload.items():
        mapping_entry = _get_fixed_row_mapping(equipment_row_slots, slot_id, "equipment")
        cell_ref = mapping_entry.get("rate")
        if not cell_ref:
            raise ValueError(
                f"Recap template map for equipment slot '{slot_id}' is missing 'rate'."
            )
        worksheet[str(cell_ref)].value = value


def _get_fixed_row_mapping(row_slots: dict[str, Any], slot_id: str, family: str) -> dict[str, Any]:
    """Return the mapped fixed-row cells for a slot-backed recap row."""
    slot_info = row_slots.get(slot_id)
    if not isinstance(slot_info, dict):
        raise ValueError(f"Recap template map is missing a {family} row mapping for slot '{slot_id}'.")
    row_mapping = slot_info.get("mapping", {})
    if not isinstance(row_mapping, dict):
        raise ValueError(f"Recap template map is missing cell mappings for {family} slot '{slot_id}'.")
    return row_mapping


def _get_fixed_row_label_cell(slot_info: dict[str, Any], family: str) -> str:
    """Return the label cell for a fixed recap row, deriving it from the mapped row when needed."""
    explicit_cell = str(slot_info.get("label_cell") or "").strip()
    if explicit_cell:
        return explicit_cell

    row_mapping = slot_info.get("mapping", {})
    if not isinstance(row_mapping, dict):
        slot_id = str(slot_info.get("slot_id") or "").strip() or "unknown"
        raise ValueError(f"Recap template map is missing cell mappings for {family} slot '{slot_id}'.")

    for key in ("st_hours", "ot_hours", "dt_hours", "st_rate", "ot_rate", "dt_rate", "hours_qty", "rate"):
        cell_ref = str(row_mapping.get(key) or "").strip()
        if not cell_ref:
            continue
        row_number = "".join(character for character in cell_ref if character.isdigit())
        if row_number:
            return f"A{row_number}"

    slot_id = str(slot_info.get("slot_id") or "").strip() or "unknown"
    raise ValueError(f"Recap template map is missing a label cell for {family} slot '{slot_id}'.")


def _write_list_section(
    worksheet: Worksheet,
    rows: list[dict[str, Any]],
    section_mapping: dict[str, Any],
    section_name: str,
) -> None:
    """Write a bounded variable-length list section into configured rows and columns."""
    start_row, end_row, columns = _get_section_bounds(section_name, section_mapping)
    capacity = end_row - start_row + 1
    if len(rows) > capacity:
        label = _SECTION_LABELS.get(section_name, section_name.title())
        raise ValueError(
            f"{label} section exceeds template capacity ({capacity} rows available, {len(rows)} required)."
        )

    for offset, row_values in enumerate(rows):
        row_number = start_row + offset
        for field_name, column_letter in columns.items():
            value = row_values.get(field_name)
            worksheet[f"{column_letter}{row_number}"].value = None if value in {None, ""} else value


def _write_sales_tax_area(
    worksheet: Worksheet,
    sales_tax_mapping: dict[str, Any],
    materials_section_mapping: dict[str, Any],
) -> None:
    """Write the sales-tax controls for the recap summary block."""
    cells = _resolve_sales_tax_cells(sales_tax_mapping, materials_section_mapping)

    section_label_cell = cells.get("section_label_cell", "")
    if section_label_cell:
        worksheet[section_label_cell].value = "Sales Tax"
    worksheet[cells["rate_label_cell"]].value = "Tax Rate"
    worksheet[cells["rate_input_cell"]].value = 0
    worksheet[cells["amount_label_cell"]].value = "Tax Amount"
    worksheet[cells["amount_formula_cell"]].value = (
        f'={cells["material_total_cell"]}*{cells["rate_input_cell"]}'
    )
    worksheet[cells["rate_input_cell"]].number_format = "0.00%"


def _write_summary_totals_area(
    worksheet: Worksheet,
    project_management_total: Any,
    sales_tax_mapping: dict[str, Any],
    materials_section_mapping: dict[str, Any],
    subcontractors_section_mapping: dict[str, Any],
    permits_section_mapping: dict[str, Any],
    police_section_mapping: dict[str, Any],
) -> None:
    """Rewrite the recap summary area to match the current modified workbook layout."""
    sales_tax_cells = _resolve_sales_tax_cells(sales_tax_mapping, materials_section_mapping)
    material_subtotal_cell = _get_section_total_cell(materials_section_mapping, default_column="H", default_end_row=41)
    subcontractor_subtotal_cell = _get_section_total_cell(
        subcontractors_section_mapping,
        default_column="C",
        default_end_row=50,
    )
    permits_total_cell = _get_section_total_cell(permits_section_mapping, default_column="C", default_end_row=56)
    police_total_cell = _get_section_total_cell(police_section_mapping, default_column="C", default_end_row=62)

    summary_values = {
        "E50": "SUMMARY & MARKUP",
        "F50": None,
        "G50": None,
        "H50": None,
        "E51": "Category",
        "F51": "Amount",
        "G51": "Control",
        "H51": "Value",
        "E52": "Labor Total",
        "F52": "=H23",
        "G52": "Material Markup %",
        "H52": 0,
        "E53": "Equipment Total",
        "F53": "=E42",
        "G53": "Material Markup",
        "H53": f'={material_subtotal_cell}*H52',
        "E54": "Material Total",
        "F54": f'={sales_tax_cells["material_total_cell"]}',
        "G54": "Material Total",
        "H54": f'={material_subtotal_cell}+H53',
        "E55": "Sales Tax",
        "F55": f'={sales_tax_cells["amount_formula_cell"]}',
        "G55": None,
        "H55": None,
        "E56": "Subcontractor Total",
        "F56": "=H58",
        "G56": "Subcontractor Markup %",
        "H56": 0,
        "E57": "Permits & Fees Total",
        "F57": f'={permits_total_cell}',
        "G57": "Subcontractor Markup",
        "H57": f'={subcontractor_subtotal_cell}*H56',
        "E58": "Police Detail Total",
        "F58": f'={police_total_cell}',
        "G58": "Subcontractor Total",
        "H58": f'={subcontractor_subtotal_cell}+H57',
        "E59": "Project Management",
        "F59": None if project_management_total in {None, ""} else project_management_total,
        "G59": None,
        "H59": None,
        "E60": None,
        "F60": None,
        "E61": None,
        "F61": None,
        "E62": None,
        "F62": None,
        "G62": None,
        "H62": None,
        "E63": "Grand Total",
        "F63": "=SUM(F52:F62)",
        "G63": None,
        "H63": None,
        "E64": None,
        "F64": None,
        "G64": None,
        "H64": None,
    }

    for cell_ref, value in summary_values.items():
        worksheet[cell_ref].value = value

    for cell_ref in ("F59", "F60", "F61", "F62"):
        _copy_cell_style(worksheet, "F58", cell_ref)


def _get_section_total_cell(
    section_mapping: dict[str, Any],
    *,
    default_column: str,
    default_end_row: int,
) -> str:
    """Return the subtotal/total cell directly beneath a configured list section."""
    columns = section_mapping.get("columns", {}) if isinstance(section_mapping, dict) else {}
    amount_column = str(columns.get("amount") or default_column).strip().upper() or default_column
    try:
        end_row = int(section_mapping.get("end_row", default_end_row)) if isinstance(section_mapping, dict) else default_end_row
    except (TypeError, ValueError):
        end_row = default_end_row
    return f"{amount_column}{end_row + 1}"


def _resolve_sales_tax_cells(
    sales_tax_mapping: dict[str, Any],
    materials_section_mapping: dict[str, Any],
) -> dict[str, str]:
    """Resolve the configured sales-tax block cells with sensible defaults."""
    default_material_total_cell = _get_section_total_cell(
        materials_section_mapping,
        default_column="H",
        default_end_row=41,
    )

    def _cell(mapping_key: str, fallback: str) -> str:
        value = str(sales_tax_mapping.get(mapping_key, fallback) or fallback).strip().upper()
        return value or fallback

    section_label_value = str(sales_tax_mapping.get("section_label_cell") or "").strip().upper()
    return {
        "section_label_cell": section_label_value,
        "rate_label_cell": _cell("rate_label_cell", "G60"),
        "rate_input_cell": _cell("rate_input_cell", "H60"),
        "amount_label_cell": _cell("amount_label_cell", "G61"),
        "amount_formula_cell": _cell("amount_formula_cell", "H61"),
        "material_total_cell": _cell("material_total_cell", default_material_total_cell),
    }


def _copy_cell_style(worksheet: Worksheet, source_cell_ref: str, target_cell_ref: str) -> None:
    """Copy the visual cell style from one template cell to another."""
    source_cell = worksheet[str(source_cell_ref)]
    target_cell = worksheet[str(target_cell_ref)]
    target_cell._style = copy(source_cell._style)


def _get_section_bounds(section_name: str, section_mapping: dict[str, Any]) -> tuple[int, int, dict[str, str]]:
    """Return validated row bounds and column mapping for a variable list section."""
    try:
        start_row = int(section_mapping["start_row"])
        end_row = int(section_mapping["end_row"])
        raw_columns = section_mapping["columns"]
    except KeyError as exc:
        raise ValueError(
            f"Recap template map for section '{section_name}' is missing required key '{exc.args[0]}'."
        ) from exc

    if end_row < start_row:
        raise ValueError(
            f"Recap template map for section '{section_name}' has an invalid row range {start_row}-{end_row}."
        )
    if not isinstance(raw_columns, dict) or not raw_columns:
        raise ValueError(
            f"Recap template map for section '{section_name}' must define one or more output columns."
        )

    columns = {
        str(field_name): str(column_letter).strip().upper()
        for field_name, column_letter in raw_columns.items()
        if str(column_letter).strip()
    }
    if not columns:
        raise ValueError(
            f"Recap template map for section '{section_name}' does not contain any usable column mappings."
        )

    return start_row, end_row, columns


def _is_zero(value: Any) -> bool:
    """Return True when a numeric export value is effectively zero."""
    try:
        return float(value) == 0.0
    except (TypeError, ValueError):
        return False
