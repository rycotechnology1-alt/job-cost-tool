from __future__ import annotations

import os
import sys
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKSPACE_ROOT = PROJECT_ROOT.parent
if str(WORKSPACE_ROOT) not in sys.path:
    sys.path.insert(0, str(WORKSPACE_ROOT))

os.environ["QT_QPA_PLATFORM"] = "offscreen"
os.environ["JOB_COST_TOOL_EXPORT_DEBUG"] = "1"

from PySide6.QtWidgets import QApplication, QFileDialog, QMessageBox

from app import window as app_window
from app.window import MainWindow
from core.export import excel_exporter, recap_mapper
from services import export_service


def main() -> int:
    sample_pdf = Path(
        r"c:\Apps\recap tool\Samples\JC Reports for Test\No way to export PM allocation 'Normalized record family is missing'\12 semi pass1.pdf"
    )
    output_path = Path(
        r"c:\Apps\recap tool\Samples\JC Reports for Test\No way to export PM allocation 'Normalized record family is missing'\12 semi pass1 UI TRACE Recap.xlsx"
    )

    if output_path.exists():
        output_path.unlink()

    print("module_paths")
    print(f"  app_window={app_window.__file__}")
    print(f"  export_service={export_service.__file__}")
    print(f"  recap_mapper={recap_mapper.__file__}")
    print(f"  excel_exporter={excel_exporter.__file__}")

    app = QApplication.instance() or QApplication([])
    window = MainWindow()

    original_get_save_file_name = QFileDialog.getSaveFileName
    original_information = QMessageBox.information
    original_critical = QMessageBox.critical
    original_question = QMessageBox.question

    def fake_get_save_file_name(*args, **kwargs):
        print(f"save_dialog_output={output_path}")
        return str(output_path), "Excel Workbook (*.xlsx)"

    def fake_information(parent, title, text, *args, **kwargs):
        print(f"info_dialog title={title!r} text={text!r}")
        return QMessageBox.StandardButton.Ok

    def fake_critical(parent, title, text, *args, **kwargs):
        print(f"critical_dialog title={title!r} text={text!r}")
        return QMessageBox.StandardButton.Ok

    def fake_question(parent, title, text, *args, **kwargs):
        print(f"question_dialog title={title!r} text={text!r}")
        return QMessageBox.StandardButton.Yes

    QFileDialog.getSaveFileName = fake_get_save_file_name
    QMessageBox.information = fake_information
    QMessageBox.critical = fake_critical
    QMessageBox.question = fake_question

    try:
        window._view_model.load_pdf(str(sample_pdf))
        print(f"current_pdf_path={window._view_model.current_pdf_path!r}")
        print(f"can_export={window._view_model.can_export!r}")
        print(f"blocking_issues={window._view_model.blocking_issues!r}")
        print(f"record_count={len(window._view_model.records)}")
        for record in window._view_model.records:
            print(
                "record",
                {
                    "phase_code": record.phase_code,
                    "raw_type": record.record_type,
                    "normalized_type": record.record_type_normalized,
                    "cost": record.cost,
                    "is_omitted": record.is_omitted,
                    "raw_description": record.raw_description,
                },
            )

        window._export_records()
    finally:
        QFileDialog.getSaveFileName = original_get_save_file_name
        QMessageBox.information = original_information
        QMessageBox.critical = original_critical
        QMessageBox.question = original_question
        window.close()
        app.quit()

    print(f"output_exists={output_path.exists()!r}")
    if output_path.exists():
        workbook = load_workbook(output_path)
        print(f"sheet_names={workbook.sheetnames!r}")
        worksheet_name = workbook.sheetnames[0]
        worksheet = workbook[worksheet_name]
        print(f"saved_sheet={worksheet_name!r}")
        print(f"saved_E59={worksheet['E59'].value!r}")
        print(f"saved_F59={worksheet['F59'].value!r}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
