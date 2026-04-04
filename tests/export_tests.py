"""Lightweight tests for recap export behavior."""

from __future__ import annotations

import shutil
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import core.export.recap_mapper as recap_mapper
from core.models.record import EQUIPMENT, LABOR, MATERIAL, PERMIT, POLICE_DETAIL, PROJECT_MANAGEMENT, SUBCONTRACTOR, Record
from services.export_service import export_records_to_recap
from services.validation_service import validate_records


TEST_TEMPLATE_MAP = {
    "worksheet_name": "Recap",
    "header_fields": {
        "project": {"cell": "B6"},
        "description": {"cell": "B7"},
        "prepared_by": {"cell": "B8"},
        "job_number": {"cell": "H6"},
        "date": {"cell": "H7"},
        "report_or_co_number": {"cell": "H8"},
    },
    "labor_rows": {
        "103 Journeyman": {
            "st_hours": "B14",
            "ot_hours": "C14",
            "dt_hours": "D14",
            "st_rate": "E14",
            "ot_rate": "F14",
            "dt_rate": "G14",
        },
        "104 Apprentice": {
            "st_hours": "B22",
            "ot_hours": "C22",
            "dt_hours": "D22",
            "st_rate": "E22",
            "ot_rate": "F22",
            "dt_rate": "G22",
        },
    },
    "equipment_rows": {
        "Pick-up Truck": {"hours_qty": "B32", "rate": "D32"},
        "Utility Van": {"hours_qty": "B33", "rate": "D33"},
    },
    "materials_section": {
        "start_row": 27,
        "end_row": 41,
        "columns": {"name": "G", "amount": "H"},
    },
    "subcontractors_section": {
        "start_row": 46,
        "end_row": 50,
        "columns": {"name": "A", "amount": "C"},
    },
    "permits_fees_section": {
        "start_row": 55,
        "end_row": 56,
        "columns": {"description": "A", "amount": "C"},
    },
    "police_detail_section": {
        "start_row": 61,
        "end_row": 62,
        "columns": {"description": "A", "amount": "C"},
    },
    "sales_tax_area": {
        "rate_label_cell": "G60",
        "rate_input_cell": "H60",
        "amount_label_cell": "G61",
        "amount_formula_cell": "H61",
        "material_total_cell": "H54",
    },
}

TEST_LABOR_SLOT_CONFIG = {
    "slots": [
        {"slot_id": "labor_1", "label": "103 Journeyman", "active": True},
        {"slot_id": "labor_2", "label": "104 Apprentice", "active": True},
    ],
    "classifications": ["103 Journeyman", "104 Apprentice"],
}
TEST_EQUIPMENT_SLOT_CONFIG = {
    "slots": [
        {"slot_id": "equipment_1", "label": "Pick-up Truck", "active": True},
        {"slot_id": "equipment_2", "label": "Utility Van", "active": True},
    ],
    "classifications": ["Pick-up Truck", "Utility Van"],
}
TEST_LABOR_ROW_SLOTS = {
    "labor_1": {
        "slot_id": "labor_1",
        "label": "103 Journeyman",
        "active": True,
        "template_label": "103 Journeyman",
        "mapping": TEST_TEMPLATE_MAP["labor_rows"]["103 Journeyman"],
    },
    "labor_2": {
        "slot_id": "labor_2",
        "label": "104 Apprentice",
        "active": True,
        "template_label": "104 Apprentice",
        "mapping": TEST_TEMPLATE_MAP["labor_rows"]["104 Apprentice"],
    },
}
TEST_EQUIPMENT_ROW_SLOTS = {
    "equipment_1": {
        "slot_id": "equipment_1",
        "label": "Pick-up Truck",
        "active": True,
        "template_label": "Pick-up Truck",
        "mapping": TEST_TEMPLATE_MAP["equipment_rows"]["Pick-up Truck"],
    },
    "equipment_2": {
        "slot_id": "equipment_2",
        "label": "Utility Van",
        "active": True,
        "template_label": "Utility Van",
        "mapping": TEST_TEMPLATE_MAP["equipment_rows"]["Utility Van"],
    },
}
TARGET_RATES = {
    "labor_rates": {
        "103 Journeyman": {"standard_rate": 199.5, "overtime_rate": 299.25, "double_time_rate": 399.0},
        "104 Apprentice": {"standard_rate": 150.0, "overtime_rate": 225.0, "double_time_rate": 300.0},
    },
    "equipment_rates": {
        "Pick-up Truck": {"rate": 88.0},
        "Utility Van": {"rate": 44.5},
    },
}
TEST_TMP_ROOT = Path("tests/_tmp")


class ExportWorkflowTests(unittest.TestCase):
    """Verify recap export safety and repeatability."""

    def setUp(self) -> None:
        TEST_TMP_ROOT.mkdir(parents=True, exist_ok=True)
        self.temp_path = TEST_TMP_ROOT / self._testMethodName
        shutil.rmtree(self.temp_path, ignore_errors=True)
        self.temp_path.mkdir(parents=True, exist_ok=True)

        self.template_path = self.temp_path / "template.xlsx"
        self.output_path = self.temp_path / "output.xlsx"
        self._create_template(self.template_path)

        recap_mapper._get_target_labor_classifications.cache_clear()
        recap_mapper._get_target_equipment_classifications.cache_clear()
        recap_mapper._get_active_labor_slots.cache_clear()
        recap_mapper._get_active_equipment_slots.cache_clear()
        recap_mapper._get_active_labor_slot_lookup.cache_clear()
        recap_mapper._get_active_equipment_slot_lookup.cache_clear()
        recap_mapper._get_rates.cache_clear()
        recap_mapper._get_material_section_capacity.cache_clear()

        self.recap_map_patch = patch(
            "core.export.excel_exporter.ConfigLoader.get_recap_template_map",
            return_value=TEST_TEMPLATE_MAP,
        )
        self.recap_map_mapper_patch = patch(
            "core.export.recap_mapper.ConfigLoader.get_recap_template_map",
            return_value=TEST_TEMPLATE_MAP,
        )
        self.labor_row_slots_patch = patch(
            "core.export.excel_exporter.ConfigLoader.get_labor_row_slots",
            return_value=TEST_LABOR_ROW_SLOTS,
        )
        self.equipment_row_slots_patch = patch(
            "core.export.excel_exporter.ConfigLoader.get_equipment_row_slots",
            return_value=TEST_EQUIPMENT_ROW_SLOTS,
        )
        self.labor_patch = patch(
            "core.export.recap_mapper.ConfigLoader.get_target_labor_classifications",
            return_value=TEST_LABOR_SLOT_CONFIG,
        )
        self.equipment_patch = patch(
            "core.export.recap_mapper.ConfigLoader.get_target_equipment_classifications",
            return_value=TEST_EQUIPMENT_SLOT_CONFIG,
        )
        self.rates_patch = patch(
            "core.export.recap_mapper.ConfigLoader.get_rates",
            return_value=TARGET_RATES,
        )
        self.recap_map_patch.start()
        self.recap_map_mapper_patch.start()
        self.labor_row_slots_patch.start()
        self.equipment_row_slots_patch.start()
        self.labor_patch.start()
        self.equipment_patch.start()
        self.rates_patch.start()

        self.addCleanup(self.recap_map_patch.stop)
        self.addCleanup(self.recap_map_mapper_patch.stop)
        self.addCleanup(self.labor_row_slots_patch.stop)
        self.addCleanup(self.equipment_row_slots_patch.stop)
        self.addCleanup(self.labor_patch.stop)
        self.addCleanup(self.equipment_patch.stop)
        self.addCleanup(self.rates_patch.stop)
        self.addCleanup(recap_mapper._get_target_labor_classifications.cache_clear)
        self.addCleanup(recap_mapper._get_target_equipment_classifications.cache_clear)
        self.addCleanup(recap_mapper._get_active_labor_slots.cache_clear)
        self.addCleanup(recap_mapper._get_active_equipment_slots.cache_clear)
        self.addCleanup(recap_mapper._get_active_labor_slot_lookup.cache_clear)
        self.addCleanup(recap_mapper._get_active_equipment_slot_lookup.cache_clear)
        self.addCleanup(recap_mapper._get_rates.cache_clear)
        self.addCleanup(recap_mapper._get_material_section_capacity.cache_clear)
        self.addCleanup(self._cleanup_temp_dir)

    def test_export_fails_when_blockers_exist(self) -> None:
        records = [self._labor_record(recap_classification=None)]

        with self.assertRaisesRegex(ValueError, "Export blocked until all blocking issues are resolved"):
            export_records_to_recap(records, str(self.template_path), str(self.output_path))

    def test_export_succeeds_after_correction(self) -> None:
        records = [
            self._labor_record(recap_classification="103 Journeyman", recap_slot_id="labor_1", hours=8),
            self._equipment_record(category="Pick-up Truck", recap_slot_id="equipment_1", hours=2),
            self._material_record(vendor="Vendor A", cost=100),
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["B6"].value, "Sample Project")
        self.assertEqual(worksheet["H6"].value, "JOB-100")
        self.assertEqual(worksheet["A14"].value, "103 Journeyman")
        self.assertEqual(worksheet["B14"].value, 8)
        self.assertEqual(worksheet["E14"].value, 199.5)
        self.assertEqual(worksheet["F14"].value, 299.25)
        self.assertEqual(worksheet["G14"].value, 399)
        self.assertEqual(worksheet["A32"].value, "Pick-up Truck")
        self.assertEqual(worksheet["B32"].value, 2)
        self.assertEqual(worksheet["D32"].value, 88)
        self.assertEqual(worksheet["G27"].value, "Vendor A")
        self.assertEqual(worksheet["H27"].value, 100)
        self.assertEqual(worksheet["H42"].value, "=SUM(H27:H41)")

    def test_export_uses_slot_rows_after_profile_label_rename(self) -> None:
        renamed_labor_slot_config = {
            "slots": [
                {"slot_id": "labor_1", "label": "Big Boy", "active": True},
                {"slot_id": "labor_2", "label": "104 Apprentice", "active": True},
            ],
            "classifications": ["Big Boy", "104 Apprentice"],
        }
        renamed_labor_row_slots = {
            **TEST_LABOR_ROW_SLOTS,
            "labor_1": {
                "slot_id": "labor_1",
                "label": "Big Boy",
                "active": True,
                "template_label": "103 Journeyman",
                "mapping": TEST_TEMPLATE_MAP["labor_rows"]["103 Journeyman"],
            },
        }
        renamed_rates = {
            **TARGET_RATES,
            "labor_rates": {
                "Big Boy": {"standard_rate": 210.0, "overtime_rate": 315.0, "double_time_rate": 420.0},
                "104 Apprentice": TARGET_RATES["labor_rates"]["104 Apprentice"],
            },
        }

        with patch(
            "core.export.recap_mapper.ConfigLoader.get_target_labor_classifications",
            return_value=renamed_labor_slot_config,
        ), patch(
            "core.export.recap_mapper.ConfigLoader.get_rates",
            return_value=renamed_rates,
        ), patch(
            "core.export.excel_exporter.ConfigLoader.get_labor_row_slots",
            return_value=renamed_labor_row_slots,
        ):
            recap_mapper._get_target_labor_classifications.cache_clear()
            recap_mapper._get_active_labor_slots.cache_clear()
            recap_mapper._get_active_labor_slot_lookup.cache_clear()
            recap_mapper._get_rates.cache_clear()

            records = [self._labor_record(recap_classification="Big Boy", recap_slot_id="labor_1", hours=8)]
            export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A14"].value, "Big Boy")
        self.assertEqual(worksheet["B14"].value, 8)
        self.assertEqual(worksheet["E14"].value, 210)
        self.assertEqual(worksheet["F14"].value, 315)
        self.assertEqual(worksheet["G14"].value, 420)

    def test_export_uses_active_equipment_slot_labels_after_profile_label_rename(self) -> None:
        renamed_equipment_slot_config = {
            "slots": [
                {"slot_id": "equipment_1", "label": "Big Rig", "active": True},
                {"slot_id": "equipment_2", "label": "Utility Van", "active": True},
            ],
            "classifications": ["Big Rig", "Utility Van"],
        }
        renamed_equipment_row_slots = {
            **TEST_EQUIPMENT_ROW_SLOTS,
            "equipment_1": {
                "slot_id": "equipment_1",
                "label": "Big Rig",
                "active": True,
                "template_label": "Pick-up Truck",
                "mapping": TEST_TEMPLATE_MAP["equipment_rows"]["Pick-up Truck"],
            },
        }
        renamed_rates = {
            **TARGET_RATES,
            "equipment_rates": {
                "Big Rig": {"rate": 99.0},
                "Utility Van": TARGET_RATES["equipment_rates"]["Utility Van"],
            },
        }

        with patch(
            "core.export.recap_mapper.ConfigLoader.get_target_equipment_classifications",
            return_value=renamed_equipment_slot_config,
        ), patch(
            "core.export.recap_mapper.ConfigLoader.get_rates",
            return_value=renamed_rates,
        ), patch(
            "core.export.excel_exporter.ConfigLoader.get_equipment_row_slots",
            return_value=renamed_equipment_row_slots,
        ):
            recap_mapper._get_target_equipment_classifications.cache_clear()
            recap_mapper._get_active_equipment_slots.cache_clear()
            recap_mapper._get_active_equipment_slot_lookup.cache_clear()
            recap_mapper._get_rates.cache_clear()

            records = [self._equipment_record(category="Big Rig", recap_slot_id="equipment_1", hours=2)]
            export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A32"].value, "Big Rig")
        self.assertEqual(worksheet["B32"].value, 2)
        self.assertEqual(worksheet["D32"].value, 99)

    def test_export_rewrites_summary_area_and_sales_tax_formulas(self) -> None:
        records = [
            self._material_record(vendor="Vendor A", cost=100),
            self._subcontractor_record(vendor="CJ Shaughnessy Crane", description="Raw subcontractor source text", cost=6000),
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["E50"].value, "SUMMARY & MARKUP")
        self.assertEqual(worksheet["E51"].value, "Category")
        self.assertEqual(worksheet["F51"].value, "Amount")
        self.assertEqual(worksheet["G51"].value, "Control")
        self.assertEqual(worksheet["H51"].value, "Value")
        self.assertEqual(worksheet["E52"].value, "Labor Total")
        self.assertEqual(worksheet["F52"].value, "=H23")
        self.assertEqual(worksheet["E53"].value, "Equipment Total")
        self.assertEqual(worksheet["F53"].value, "=E42")
        self.assertEqual(worksheet["E54"].value, "Material Total")
        self.assertEqual(worksheet["F54"].value, "=H54")
        self.assertEqual(worksheet["E55"].value, "Sales Tax")
        self.assertEqual(worksheet["F55"].value, "=H61")
        self.assertEqual(worksheet["E56"].value, "Subcontractor Total")
        self.assertEqual(worksheet["F56"].value, "=H58")
        self.assertEqual(worksheet["E57"].value, "Permits & Fees Total")
        self.assertEqual(worksheet["F57"].value, "=C57")
        self.assertEqual(worksheet["E58"].value, "Police Detail Total")
        self.assertEqual(worksheet["F58"].value, "=C63")
        self.assertEqual(worksheet["E59"].value, "Project Management")
        self.assertIsNone(worksheet["F59"].value)
        self.assertEqual(worksheet["E63"].value, "Grand Total")
        self.assertEqual(worksheet["F63"].value, "=SUM(F52:F62)")
        self.assertEqual(worksheet["G52"].value, "Material Markup %")
        self.assertEqual(worksheet["H52"].value, 0)
        self.assertEqual(worksheet["G53"].value, "Material Markup")
        self.assertEqual(worksheet["H53"].value, "=H42*H52")
        self.assertEqual(worksheet["G54"].value, "Material Total")
        self.assertEqual(worksheet["H54"].value, "=H42+H53")
        self.assertEqual(worksheet["G56"].value, "Subcontractor Markup %")
        self.assertEqual(worksheet["H56"].value, 0)
        self.assertEqual(worksheet["G57"].value, "Subcontractor Markup")
        self.assertEqual(worksheet["H57"].value, "=C51*H56")
        self.assertEqual(worksheet["G58"].value, "Subcontractor Total")
        self.assertEqual(worksheet["H58"].value, "=C51+H57")
        self.assertEqual(worksheet["G60"].value, "Tax Rate")
        self.assertEqual(worksheet["H60"].value, 0)
        self.assertEqual(worksheet["H60"].number_format, "0.00%")
        self.assertEqual(worksheet["G61"].value, "Tax Amount")
        self.assertEqual(worksheet["H61"].value, "=H54*H60")
        self.assertIsNone(worksheet["G63"].value)
        self.assertIsNone(worksheet["H65"].value)

    def test_export_succeeds_with_actual_modified_default_template(self) -> None:
        actual_template = Path("profiles/default/recap_template.xlsx")
        records = [
            self._material_record(vendor=f"Vendor {index}", cost=10 + index)
            for index in range(10)
        ] + [
            self._subcontractor_record(vendor="CJ Shaughnessy Crane", description="Raw subcontractor source text", cost=6000),
        ]

        export_records_to_recap(records, str(actual_template), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["H6"].value, "JOB-100")
        self.assertEqual(worksheet["G27"].value, "Vendor 0")
        self.assertEqual(worksheet["H27"].value, 10)
        self.assertEqual(worksheet["G34"].value, "Vendor 7")
        self.assertEqual(worksheet["H34"].value, 17)
        self.assertEqual(worksheet["G27"].style_id, worksheet["G34"].style_id)
        self.assertEqual(worksheet["H27"].style_id, worksheet["H34"].style_id)
        self.assertEqual(worksheet["A46"].value, "CJ Shaughnessy Crane")
        self.assertEqual(worksheet["C46"].value, 6000)
        self.assertEqual(worksheet["F63"].value, "=SUM(F52:F62)")

    def test_export_leaves_subcontractor_description_cells_blank(self) -> None:
        records = [
            self._subcontractor_record(vendor="CJ Shaughnessy Crane", description="Raw subcontractor source text", cost=6000)
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A46"].value, "CJ Shaughnessy Crane")
        self.assertIsNone(worksheet["B46"].value)
        self.assertEqual(worksheet["C46"].value, 6000)

    def test_export_routes_phase_50_point_1_permit_records_to_permits_fees_section(self) -> None:
        records = [
            self._permit_record(
                phase_code="50 .1",
                description="408 Bank of America BOA 3-2-26 / TR# 8 / 0 / APCo: 2 BOA 1446 3-2-26",
                cost=1293.39,
                vendor_name="Bank of America BOA",
            )
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A55"].value, "Bank of America BOA")
        self.assertEqual(worksheet["C55"].value, 1293.39)
        self.assertIsNone(worksheet["G27"].value)
        self.assertIsNone(worksheet["H27"].value)

    def test_export_uses_raw_description_for_permit_rows_only_when_vendor_is_missing(self) -> None:
        records = [
            self._permit_record(
                phase_code="50 .1",
                description="City Permit 12345",
                cost=250.0,
                vendor_name=None,
            )
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A55"].value, "City Permit 12345")
        self.assertEqual(worksheet["C55"].value, 250)

    def test_export_routes_phase_50_point_2_police_records_to_police_detail_section(self) -> None:
        records = [
            self._police_detail_record(
                phase_code="50 .2",
                description="22714 Project Flagging LLC 63164 / TR# 163 / 0 / APCo: 1 Flagging - 220108",
                cost=922.50,
                vendor_name="Project Flagging LLC",
            )
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A61"].value, "Project Flagging LLC")
        self.assertEqual(worksheet["C61"].value, 922.5)
        self.assertIsNone(worksheet["G27"].value)
        self.assertIsNone(worksheet["H27"].value)

    def test_export_uses_raw_description_for_police_rows_only_when_vendor_is_missing(self) -> None:
        records = [
            self._police_detail_record(
                phase_code="50 .2",
                description="Police Detail Ticket 7788",
                cost=175.0,
                vendor_name=None,
            )
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["A61"].value, "Police Detail Ticket 7788")
        self.assertEqual(worksheet["C61"].value, 175)

    def test_export_writes_project_management_total_into_summary(self) -> None:
        records = [
            self._project_management_record(cost=20000.0),
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["E59"].value, "Project Management")
        self.assertEqual(worksheet["F59"].value, 20000)
        self.assertEqual(worksheet["F63"].value, "=SUM(F52:F62)")
        self.assertEqual(worksheet["F58"].style_id, worksheet["F59"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F60"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F61"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F62"].style_id)
        self.assertIsNone(worksheet["G27"].value)
        self.assertIsNone(worksheet["A55"].value)

    def test_export_sums_multiple_included_project_management_records_into_summary(self) -> None:
        records = [
            self._project_management_record(cost=20000.0),
            self._project_management_record(cost=5000.25),
            self._project_management_record(cost=750.75, is_omitted=True),
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["E59"].value, "Project Management")
        self.assertEqual(worksheet["F59"].value, 25000.25)
        self.assertEqual(worksheet["F63"].value, "=SUM(F52:F62)")

    def test_export_writes_project_management_total_into_actual_modified_default_template(self) -> None:
        actual_template = Path("profiles/default/recap_template.xlsx")
        records = [
            self._project_management_record(cost=20000.0),
            self._material_record(vendor="Vendor A", cost=100),
        ]

        export_records_to_recap(records, str(actual_template), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["E59"].value, "Project Management")
        self.assertEqual(worksheet["F59"].value, 20000)
        self.assertEqual(worksheet["F63"].value, "=SUM(F52:F62)")
        self.assertEqual(worksheet["F58"].style_id, worksheet["F59"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F60"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F61"].style_id)
        self.assertEqual(worksheet["F58"].style_id, worksheet["F62"].style_id)


    def test_export_collapses_material_vendor_overflow_into_additional_vendors(self) -> None:
        records = [
            self._material_record(vendor=f"Vendor {index}", cost=10 + index)
            for index in range(16)
        ]

        export_records_to_recap(records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["G27"].value, "Vendor 0")
        self.assertEqual(worksheet["H27"].value, 10)
        self.assertEqual(worksheet["G40"].value, "Vendor 13")
        self.assertEqual(worksheet["H40"].value, 23)
        self.assertEqual(worksheet["G41"].value, "Additional Vendors")
        self.assertEqual(worksheet["H41"].value, 49)
        self.assertEqual(worksheet["H42"].value, "=SUM(H27:H41)")

    def test_material_overflow_uses_current_template_capacity(self) -> None:
        smaller_template_map = {
            **TEST_TEMPLATE_MAP,
            "materials_section": {
                **TEST_TEMPLATE_MAP["materials_section"],
                "end_row": 29,
            },
        }
        records = [
            self._material_record(vendor=f"Vendor {index}", cost=10 + index)
            for index in range(4)
        ]

        with patch(
            "core.export.recap_mapper.ConfigLoader.get_recap_template_map",
            return_value=smaller_template_map,
        ):
            recap_mapper._get_material_section_capacity.cache_clear()
            payload = recap_mapper.build_recap_payload(records)

        self.assertEqual(
            payload["materials"],
            [
                {"name": "Vendor 0", "amount": 10},
                {"name": "Vendor 1", "amount": 11},
                {"name": "Additional Vendors", "amount": 25},
            ],
        )

    def test_material_overflow_preserves_vendors_by_first_appearance_order(self) -> None:
        smaller_template_map = {
            **TEST_TEMPLATE_MAP,
            "materials_section": {
                **TEST_TEMPLATE_MAP["materials_section"],
                "end_row": 29,
            },
        }
        records = [
            self._material_record(vendor="Vendor C", cost=30),
            self._material_record(vendor="Vendor A", cost=10),
            self._material_record(vendor="Vendor C", cost=5),
            self._material_record(vendor="Vendor B", cost=20),
            self._material_record(vendor="Vendor D", cost=40),
        ]

        with patch(
            "core.export.recap_mapper.ConfigLoader.get_recap_template_map",
            return_value=smaller_template_map,
        ):
            recap_mapper._get_material_section_capacity.cache_clear()
            payload = recap_mapper.build_recap_payload(records)

        self.assertEqual(
            payload["materials"],
            [
                {"name": "Vendor C", "amount": 35},
                {"name": "Vendor A", "amount": 10},
                {"name": "Additional Vendors", "amount": 60},
            ],
        )

    def test_export_fails_if_rate_target_cell_mapping_is_missing(self) -> None:
        broken_labor_row_slots = {
            **TEST_LABOR_ROW_SLOTS,
            "labor_1": {
                **TEST_LABOR_ROW_SLOTS["labor_1"],
                "mapping": {
                    key: value
                    for key, value in TEST_LABOR_ROW_SLOTS["labor_1"]["mapping"].items()
                    if key != "st_rate"
                },
            },
        }

        with patch(
            "core.export.excel_exporter.ConfigLoader.get_labor_row_slots",
            return_value=broken_labor_row_slots,
        ):
            records = [self._labor_record(recap_classification="103 Journeyman", recap_slot_id="labor_1", hours=8)]
            with self.assertRaisesRegex(ValueError, "missing 'st_rate'"):
                export_records_to_recap(records, str(self.template_path), str(self.output_path))

    def test_validation_blocks_missing_labor_hour_type_before_export(self) -> None:
        records = [
            self._labor_record(
                recap_classification="103 Journeyman",
                recap_slot_id="labor_1",
                hours=-4,
                hour_type=None,
            )
        ]

        validated_records, blocking_issues = validate_records(records)

        self.assertIn(
            "Record on page 1 (phase 20, labor): Labor hour type is missing for export.",
            blocking_issues,
        )
        self.assertIn(
            "BLOCKING: Labor hour type is missing for export.",
            validated_records[0].warnings,
        )

    def test_export_blocks_missing_labor_hour_type_before_recap_build(self) -> None:
        records = [
            self._labor_record(
                recap_classification="103 Journeyman",
                recap_slot_id="labor_1",
                hours=-4,
                hour_type=None,
            )
        ]

        with patch("services.export_service.build_recap_payload") as build_payload_mock:
            with self.assertRaisesRegex(ValueError, "Labor hour type is missing for export"):
                export_records_to_recap(records, str(self.template_path), str(self.output_path))

        build_payload_mock.assert_not_called()


    def test_export_fails_if_template_missing(self) -> None:
        records = [self._labor_record(recap_classification="103 Journeyman", recap_slot_id="labor_1")]
        missing_template = self.temp_path / "missing-template.xlsx"

        with self.assertRaisesRegex(FileNotFoundError, "Recap template workbook was not found"):
            export_records_to_recap(records, str(missing_template), str(self.output_path))

    def test_export_clears_previous_data_when_reused(self) -> None:
        first_records = [
            self._material_record(vendor="Vendor A", cost=100),
            self._material_record(vendor="Vendor B", cost=200),
        ]
        second_records = [
            self._material_record(vendor="Vendor C", cost=300),
        ]

        export_records_to_recap(first_records, str(self.template_path), str(self.output_path))
        export_records_to_recap(second_records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertEqual(worksheet["G27"].value, "Vendor C")
        self.assertEqual(worksheet["H27"].value, 300)
        self.assertIsNone(worksheet["G28"].value)
        self.assertIsNone(worksheet["H28"].value)
        self.assertEqual(worksheet["H42"].value, "=SUM(H27:H41)")

    def test_omitted_record_does_not_block_or_export(self) -> None:
        records = [
            self._labor_record(recap_classification=None, is_omitted=True),
            self._material_record(vendor="Vendor A", cost=100),
        ]

        validated_records, blocking_issues = validate_records(records)
        self.assertEqual(blocking_issues, [])

        export_records_to_recap(validated_records, str(self.template_path), str(self.output_path))

        worksheet = load_workbook(self.output_path)["Recap"]
        self.assertIsNone(worksheet["B14"].value)
        self.assertEqual(worksheet["G27"].value, "Vendor A")
        self.assertEqual(worksheet["H27"].value, 100)

    def _cleanup_temp_dir(self) -> None:
        shutil.rmtree(self.temp_path, ignore_errors=True)

    def _create_template(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Recap"

        for cell, value in {
            "A6": "Project",
            "A7": "Description",
            "A8": "Prepared By",
            "G6": "Job Number",
            "G7": "Date",
            "G8": "Report / CO #",
            "A25": "EQUIPMENT",
            "A26": "Category",
            "B26": "Hours / Qty",
            "C26": "Unit",
            "D26": "Rate",
            "E26": "Total",
            "G25": "MATERIALS",
            "G26": "Vendor",
            "H26": "Amount",
            "A42": "Equipment Total",
            "E42": "=SUM(E27:E41)",
            "G42": "Material Subtotal",
            "H42": "=SUM(H27:H41)",
            "A44": "SUBCONTRACTORS",
            "A45": "Subcontractor Name",
            "C45": "Amount",
            "A51": "Subcontractor Subtotal",
            "C51": "=SUM(C46:C50)",
            "A53": "PERMITS & FEES",
            "A54": "Description",
            "C54": "Amount",
            "A57": "Permits & Fees Total",
            "C57": "=SUM(C55:C56)",
            "A59": "POLICE DETAIL",
            "A60": "Description",
            "C60": "Amount",
            "A63": "Police Detail Total",
            "C63": "=SUM(C61:C62)",
            "H23": "=SUM(H12:H22)",
        }.items():
            worksheet[cell] = value

        workbook.save(path)

    def _labor_record(
        self,
        recap_classification: str | None,
        hours: float = 8,
        is_omitted: bool = False,
        recap_slot_id: str | None = None,
        hour_type: str | None = "ST",
    ) -> Record:
        return Record(
            record_type=LABOR,
            phase_code="20",
            raw_description="Labor line",
            cost=100,
            hours=hours,
            hour_type=hour_type,
            union_code="103",
            labor_class_raw="J",
            labor_class_normalized="J",
            vendor_name=None,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Labor source",
            record_type_normalized=LABOR,
            recap_labor_slot_id=recap_slot_id,
            recap_labor_classification=recap_classification,
            vendor_name_normalized=None,
            is_omitted=is_omitted,
        )

    def _equipment_record(self, category: str, hours: float = 2, recap_slot_id: str | None = None) -> Record:
        return Record(
            record_type=EQUIPMENT,
            phase_code="31",
            raw_description="Equipment line",
            cost=50,
            hours=hours,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description="Pickup truck",
            equipment_category=category,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Equipment source",
            record_type_normalized=EQUIPMENT,
            recap_labor_classification=None,
            recap_equipment_slot_id=recap_slot_id,
            vendor_name_normalized=None,
        )

    def _material_record(self, vendor: str, cost: float) -> Record:
        return Record(
            record_type=MATERIAL,
            phase_code="50",
            raw_description="Material line",
            cost=cost,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=vendor,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Material source",
            record_type_normalized=MATERIAL,
            recap_labor_classification=None,
            vendor_name_normalized=vendor,
        )

    def _project_management_record(self, cost: float, is_omitted: bool = False) -> Record:
        return Record(
            record_type=PROJECT_MANAGEMENT,
            phase_code="25",
            raw_description="Bugeted PM Allocation",
            cost=cost,
            hours=0.0,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Project management source",
            record_type_normalized=PROJECT_MANAGEMENT,
            recap_labor_classification=None,
            vendor_name_normalized=None,
            is_omitted=is_omitted,
        )

    def _subcontractor_record(self, vendor: str, description: str, cost: float) -> Record:
        return Record(
            record_type=SUBCONTRACTOR,
            phase_code="40",
            raw_description=description,
            cost=cost,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=vendor,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Subcontractor source",
            record_type_normalized=SUBCONTRACTOR,
            recap_labor_classification=None,
            vendor_name_normalized=vendor,
        )

    def _permit_record(self, phase_code: str, description: str, cost: float, vendor_name: str | None = "Bank of America") -> Record:
        return Record(
            record_type=PERMIT,
            phase_code=phase_code,
            raw_description=description,
            cost=cost,
            hours=0.0,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=vendor_name,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Permit source",
            record_type_normalized=PERMIT,
            recap_labor_classification=None,
            vendor_name_normalized=vendor_name,
        )

    def _police_detail_record(self, phase_code: str, description: str, cost: float, vendor_name: str | None = "Project Flagging LLC") -> Record:
        return Record(
            record_type=POLICE_DETAIL,
            phase_code=phase_code,
            raw_description=description,
            cost=cost,
            hours=0.0,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=vendor_name,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Police detail source",
            record_type_normalized=POLICE_DETAIL,
            recap_labor_classification=None,
            vendor_name_normalized=vendor_name,
        )

if __name__ == "__main__":
    unittest.main()
