"""Focused Postgres integration coverage for the compatibility lineage store."""

from __future__ import annotations

import json
import os
import shutil
import unittest
from datetime import datetime, timezone
from pathlib import Path
from uuid import uuid4
from unittest.mock import patch

import psycopg
from psycopg import sql
from openpyxl import Workbook, load_workbook

from api.dependencies import build_runtime
from core.config import ConfigLoader, ProfileManager
from core.models import MATERIAL, PendingRecordEdit, Record
from core.models.lineage import User
from infrastructure.persistence import PostgresLineageStore, SqliteLineageStore
from infrastructure.persistence.sqlite_to_postgres_import import import_sqlite_lineage_to_postgres
from services.profile_execution_compatibility_adapter import ProfileExecutionCompatibilityAdapter
from services.profile_authoring_errors import ProfileAuthoringPersistenceConflictError
from services.profile_authoring_service import ProfileAuthoringService
from services.processing_run_service import ProcessingRunService
from services.request_context import RequestContext
from services.review_session_service import ReviewSessionService
from services.trusted_profile_authoring_repository import TrustedProfileAuthoringRepository
from services.trusted_profile_provisioning_service import TrustedProfileProvisioningService


TEST_ROOT = Path("tests/_postgres_lineage_tmp")


class PostgresLineageStoreTests(unittest.TestCase):
    """Verify compatibility Postgres persistence preserves current lineage behavior."""

    def setUp(self) -> None:
        self.postgres_admin_url = os.environ.get("JOB_COST_API_POSTGRES_ADMIN_URL", "").strip()
        self.postgres_pooled_url = os.environ.get("JOB_COST_API_POSTGRES_POOLED_URL", "").strip()
        if not self.postgres_admin_url or not self.postgres_pooled_url:
            raise unittest.SkipTest("Postgres integration tests require JOB_COST_API_POSTGRES_ADMIN_URL and JOB_COST_API_POSTGRES_POOLED_URL.")

        ConfigLoader.clear_runtime_caches()
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        (TEST_ROOT / "profiles" / "default").mkdir(parents=True, exist_ok=True)
        (TEST_ROOT / "legacy_config").mkdir(parents=True, exist_ok=True)
        self.settings_path = TEST_ROOT / "app_settings.json"
        self.source_document_path = TEST_ROOT / "sample_report.pdf"
        self.source_document_path.write_bytes(b"sample pdf bytes")
        self.created_at = datetime(2026, 4, 18, 12, 0, tzinfo=timezone.utc)
        self.schema_names: list[str] = []

        self._write_profile_bundle()
        self._write_json(self.settings_path, {"active_profile": "default"})
        self._write_json(TEST_ROOT / "legacy_config" / "phase_catalog.json", {"phases": []})

        self.profile_manager = ProfileManager(
            profiles_root=TEST_ROOT / "profiles",
            legacy_config_root=TEST_ROOT / "legacy_config",
        )
        self.lineage_store = self._create_postgres_store()
        (
            self.repository,
            self.trusted_profile_provisioning_service,
            self.profile_execution_compatibility_adapter,
            self.profile_authoring_service,
            self.processing_run_service,
            self.review_session_service,
        ) = self._build_service_stack(self.lineage_store)

    def tearDown(self) -> None:
        if hasattr(self, "lineage_store"):
            self.lineage_store.close()
        for schema_name in reversed(getattr(self, "schema_names", [])):
            self._drop_schema(schema_name)
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        ConfigLoader.clear_runtime_caches()

    def test_build_runtime_can_select_postgres_store_via_configuration(self) -> None:
        schema_name = self._new_schema_name("runtime")

        runtime, owns_lineage_store = build_runtime(
            database_provider="postgres",
            postgres_admin_url=self.postgres_admin_url,
            postgres_pooled_url=self.postgres_pooled_url,
            postgres_schema=schema_name,
            profile_manager=self.profile_manager,
            upload_root=TEST_ROOT / "runtime" / "uploads",
            export_root=TEST_ROOT / "runtime" / "exports",
            engine_version="engine-1",
            now_provider=lambda: self.created_at,
        )
        try:
            self.assertTrue(owns_lineage_store)
            self.assertIsInstance(runtime.lineage_store, PostgresLineageStore)
        finally:
            runtime.lineage_store.close()

    def test_postgres_purges_workflow_rows_after_export_and_keeps_retained_export_metadata(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id

        updated_state = self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )

        export_output = TEST_ROOT / "exports" / "revision-1.xlsx"
        export_output.parent.mkdir(parents=True, exist_ok=True)
        export_result = self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=1,
            output_path=export_output,
        )

        persisted_run_records = self.lineage_store.list_run_records(processing_run_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(
            updated_state.review_session.review_session_id
        )
        persisted_artifacts = self.lineage_store.list_export_artifacts(
            updated_state.review_session.review_session_id
        )
        worksheet = load_workbook(export_output)["Recap"]

        self.assertEqual(updated_state.review_session.current_revision, 1)
        self.assertEqual(updated_state.session_revision, 1)
        self.assertEqual(export_result.export_artifact.session_revision, 1)
        self.assertEqual(persisted_run_records, [])
        self.assertEqual(persisted_edits, [])
        self.assertEqual(persisted_artifacts[0].session_revision, 1)
        self.assertEqual(worksheet["G27"].value, "Vendor B")

    def test_postgres_store_rejects_stale_review_revision_at_persistence_layer(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        self.lineage_store.ensure_user(
            User(
                user_id="user-1",
                organization_id="org-default",
                email="user1@example.com",
                display_name="User One",
                auth_subject="auth-user-1",
                is_active=True,
                created_at=self.created_at,
            )
        )
        self.lineage_store.ensure_user(
            User(
                user_id="user-2",
                organization_id="org-default",
                email="user2@example.com",
                display_name="User Two",
                auth_subject="auth-user-2",
                is_active=True,
                created_at=self.created_at,
            )
        )

        first_state = self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
            expected_current_revision=0,
            request_context=RequestContext(
                organization_id="org-default",
                user_id="user-1",
                role="member",
            ),
        )

        with self.assertRaises(ProfileAuthoringPersistenceConflictError):
            self.review_session_service.apply_review_edits(
                processing_run_id,
                [
                    PendingRecordEdit(
                        record_key="record-0",
                        changed_fields={"vendor_name_normalized": "Vendor C"},
                    )
                ],
                expected_current_revision=0,
                request_context=RequestContext(
                    organization_id="org-default",
                    user_id="user-2",
                    role="member",
                ),
            )

        persisted_session = self.lineage_store.get_review_session(first_state.review_session.review_session_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(first_state.review_session.review_session_id)

        self.assertEqual(persisted_session.current_revision, 1)
        self.assertEqual(len(persisted_edits), 1)

    def test_postgres_review_revision_advance_rolls_back_when_edit_insert_fails(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        review_session = self.review_session_service.open_review_session(processing_run_id).review_session
        self.lineage_store._connection.execute(
            """
            CREATE OR REPLACE FUNCTION reviewed_record_edits_block_insert()
            RETURNS trigger
            LANGUAGE plpgsql
            AS $$
            BEGIN
                RAISE EXCEPTION 'block review edit insert';
            END;
            $$;
            """
        )
        self.lineage_store._connection.execute(
            """
            CREATE TRIGGER reviewed_record_edits_block_insert
            BEFORE INSERT ON reviewed_record_edits
            FOR EACH ROW
            EXECUTE FUNCTION reviewed_record_edits_block_insert();
            """
        )
        self.lineage_store._connection.commit()

        with self.assertRaisesRegex(psycopg.Error, "block review edit insert"):
            self.review_session_service.apply_review_edits(
                processing_run_id,
                [
                    PendingRecordEdit(
                        record_key="record-0",
                        changed_fields={"vendor_name_normalized": "Vendor B"},
                    )
                ],
            )

        persisted_session = self.lineage_store.get_review_session(review_session.review_session_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(review_session.review_session_id)

        self.assertEqual(persisted_session.current_revision, 0)
        self.assertEqual(persisted_edits, [])

    def test_postgres_draft_revision_increments_on_real_save(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"

        self.trusted_profile_provisioning_service.get_trusted_profile(trusted_profile_id)
        draft = self.repository.create_open_draft("org-default", trusted_profile_id)
        updated_bundle = json.loads(json.dumps(draft.bundle_payload))
        updated_bundle["behavioral_bundle"]["review_rules"]["default_omit_rules"] = [
            {"phase_code": "20", "phase_name": "Labor"}
        ]
        saved_draft = self.repository.save_draft_bundle(
            draft.organization_id,
            draft.trusted_profile_draft_id,
            updated_bundle,
            expected_draft_revision=draft.draft_revision,
        )
        stored_draft = self.lineage_store.get_trusted_profile_draft(saved_draft.trusted_profile_draft_id)

        self.assertEqual(draft.draft_revision, 1)
        self.assertEqual(saved_draft.draft_revision, 2)
        self.assertEqual(stored_draft.draft_revision, 2)
        self.assertEqual(
            stored_draft.bundle_payload["behavioral_bundle"]["review_rules"]["default_omit_rules"],
            [{"phase_code": "20", "phase_name": "Labor"}],
        )

    def test_postgres_store_rejects_stale_draft_save_at_persistence_layer(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"

        self.trusted_profile_provisioning_service.get_trusted_profile(trusted_profile_id)
        stale_draft = self.repository.create_open_draft("org-default", trusted_profile_id)
        first_update = json.loads(json.dumps(stale_draft.bundle_payload))
        first_update["behavioral_bundle"]["review_rules"]["default_omit_rules"] = [
            {"phase_code": "20", "phase_name": "Labor"}
        ]
        self.repository.save_draft_bundle(
            stale_draft.organization_id,
            stale_draft.trusted_profile_draft_id,
            first_update,
            expected_draft_revision=stale_draft.draft_revision,
        )

        second_update = json.loads(json.dumps(stale_draft.bundle_payload))
        second_update["behavioral_bundle"]["review_rules"]["default_omit_rules"] = [
            {"phase_code": "30", "phase_name": "Equipment"}
        ]

        with patch.object(self.repository, "get_draft", return_value=stale_draft):
            with self.assertRaises(ProfileAuthoringPersistenceConflictError):
                self.repository.save_draft_bundle(
                    stale_draft.organization_id,
                    stale_draft.trusted_profile_draft_id,
                    second_update,
                    expected_draft_revision=stale_draft.draft_revision,
                )

    def test_postgres_publish_rejects_stale_expected_draft_revision(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"
        draft_state = self.profile_authoring_service.create_or_open_draft(trusted_profile_id)

        updated_state = self.profile_authoring_service.update_default_omit_rules(
            draft_state.trusted_profile_draft_id,
            [{"phase_code": "20", "phase_name": "Labor"}],
            expected_draft_revision=draft_state.draft_revision,
        )

        with self.assertRaises(ProfileAuthoringPersistenceConflictError):
            self.profile_authoring_service.publish_draft(
                updated_state.trusted_profile_draft_id,
                expected_draft_revision=draft_state.draft_revision,
                request_context=RequestContext(
                    organization_id="org-default",
                    user_id="user-1",
                    role="member",
                ),
            )

    def test_postgres_publish_rejects_stale_second_publish_after_draft_is_already_claimed(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"
        draft_state = self.profile_authoring_service.create_or_open_draft(trusted_profile_id)
        updated_state = self.profile_authoring_service.update_default_omit_rules(
            draft_state.trusted_profile_draft_id,
            [{"phase_code": "20", "phase_name": "Labor"}],
            expected_draft_revision=draft_state.draft_revision,
        )
        stale_draft = self.repository.get_draft("org-default", updated_state.trusted_profile_draft_id)

        published_version = self.repository.publish_draft(
            "org-default",
            updated_state.trusted_profile_draft_id,
            expected_draft_revision=updated_state.draft_revision,
        )

        with patch.object(self.repository, "get_draft", return_value=stale_draft):
            with self.assertRaises(ProfileAuthoringPersistenceConflictError):
                self.repository.publish_draft(
                    "org-default",
                    updated_state.trusted_profile_draft_id,
                    expected_draft_revision=updated_state.draft_revision,
                )

        versions = self.lineage_store.list_trusted_profile_versions(trusted_profile_id)
        self.assertEqual(
            published_version.trusted_profile_version_id,
            self.lineage_store.get_trusted_profile(trusted_profile_id).current_published_version_id,
        )
        self.assertEqual(len(versions), 2)

    def test_postgres_publish_rolls_back_version_and_pointer_when_draft_delete_fails(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"
        original_detail = self.profile_authoring_service.get_profile_detail(trusted_profile_id)
        draft_state = self.profile_authoring_service.create_or_open_draft(trusted_profile_id)
        updated_state = self.profile_authoring_service.update_default_omit_rules(
            draft_state.trusted_profile_draft_id,
            [{"phase_code": "20", "phase_name": "Labor"}],
            expected_draft_revision=draft_state.draft_revision,
        )

        self.lineage_store._connection.execute(
            """
            CREATE OR REPLACE FUNCTION trusted_profile_drafts_block_delete()
            RETURNS trigger
            LANGUAGE plpgsql
            AS $$
            BEGIN
                RAISE EXCEPTION 'block publish draft delete';
            END;
            $$;
            """
        )
        self.lineage_store._connection.execute(
            """
            CREATE TRIGGER trusted_profile_drafts_block_delete
            BEFORE DELETE ON trusted_profile_drafts
            FOR EACH ROW
            EXECUTE FUNCTION trusted_profile_drafts_block_delete();
            """
        )
        self.lineage_store._connection.commit()

        with self.assertRaises(psycopg.Error):
            self.profile_authoring_service.publish_draft(
                updated_state.trusted_profile_draft_id,
                expected_draft_revision=updated_state.draft_revision,
                request_context=RequestContext(
                    organization_id="org-default",
                    user_id="user-1",
                    role="member",
                ),
            )

        versions = self.lineage_store.list_trusted_profile_versions(trusted_profile_id)
        current_profile = self.lineage_store.get_trusted_profile(trusted_profile_id)
        persisted_draft = self.repository.get_draft("org-default", updated_state.trusted_profile_draft_id)

        self.assertEqual(len(versions), 1)
        self.assertEqual(
            current_profile.current_published_version_id,
            original_detail.current_published_version_id,
        )
        self.assertEqual(
            persisted_draft.trusted_profile_draft_id,
            updated_state.trusted_profile_draft_id,
        )

    def test_postgres_forward_draft_revision_migration_is_schema_safe(self) -> None:
        legacy_schema_a = self._new_schema_name("legacy_a")
        legacy_schema_b = self._new_schema_name("legacy_b")
        migration_sql = Path("infrastructure/persistence/postgres_migrations/0003_phase5_draft_revision.sql").read_text(
            encoding="utf-8"
        )

        with psycopg.connect(self.postgres_admin_url) as connection:
            for schema_name in (legacy_schema_a, legacy_schema_b):
                connection.execute(
                    sql.SQL("CREATE SCHEMA IF NOT EXISTS {}").format(sql.Identifier(schema_name))
                )
                connection.execute(
                    sql.SQL("SET search_path TO {}, public").format(sql.Identifier(schema_name))
                )
                connection.execute(
                    """
                    CREATE TABLE IF NOT EXISTS trusted_profile_drafts (
                        trusted_profile_draft_id TEXT PRIMARY KEY,
                        organization_id TEXT NOT NULL,
                        trusted_profile_id TEXT NOT NULL,
                        base_trusted_profile_version_id TEXT,
                        bundle_json TEXT NOT NULL,
                        content_hash TEXT NOT NULL,
                        template_artifact_id TEXT,
                        template_artifact_ref TEXT,
                        template_file_hash TEXT,
                        status TEXT NOT NULL DEFAULT 'open' CHECK (status = 'open'),
                        created_by_user_id TEXT,
                        created_at TEXT NOT NULL,
                        updated_at TEXT NOT NULL,
                        UNIQUE (trusted_profile_id)
                    )
                    """
                )

            connection.execute(sql.SQL("SET search_path TO {}, public").format(sql.Identifier(legacy_schema_a)))
            connection.execute(migration_sql)
            connection.execute(sql.SQL("SET search_path TO {}, public").format(sql.Identifier(legacy_schema_b)))
            connection.execute(migration_sql)

            column_exists = connection.execute(
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema = %s
                  AND table_name = 'trusted_profile_drafts'
                  AND column_name = 'draft_revision'
                """,
                (legacy_schema_b,),
            ).fetchone()
            constraint_exists = connection.execute(
                """
                SELECT 1
                FROM pg_constraint c
                JOIN pg_class rel ON rel.oid = c.conrelid
                JOIN pg_namespace ns ON ns.oid = rel.relnamespace
                WHERE ns.nspname = %s
                  AND rel.relname = 'trusted_profile_drafts'
                  AND c.conname = 'ck_trusted_profile_drafts_draft_revision'
                """,
                (legacy_schema_b,),
            ).fetchone()
            connection.execute(sql.SQL("SET search_path TO {}, public").format(sql.Identifier(legacy_schema_b)))
            connection.execute(
                """
                INSERT INTO trusted_profile_drafts (
                    trusted_profile_draft_id,
                    organization_id,
                    trusted_profile_id,
                    base_trusted_profile_version_id,
                    bundle_json,
                    content_hash,
                    template_artifact_id,
                    template_artifact_ref,
                    template_file_hash,
                    status,
                    created_by_user_id,
                    created_at,
                    updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    "draft-1",
                    "org-default",
                    "trusted-profile:org-default:default",
                    None,
                    "{}",
                    "content-hash",
                    None,
                    None,
                    None,
                    "open",
                    None,
                    self.created_at.isoformat(),
                    self.created_at.isoformat(),
                ),
            )
            revision = connection.execute(
                """
                SELECT draft_revision
                FROM trusted_profile_drafts
                WHERE trusted_profile_draft_id = %s
                """,
                ("draft-1",),
            ).fetchone()

        self.assertIsNotNone(column_exists)
        self.assertIsNotNone(constraint_exists)
        self.assertEqual(revision[0], 1)

    def test_postgres_preserves_draft_publish_immutability_and_processing_snapshot_behavior(self) -> None:
        trusted_profile_id = "trusted-profile:org-default:default"
        original_detail = self.profile_authoring_service.get_profile_detail(trusted_profile_id)
        draft_state = self.profile_authoring_service.create_or_open_draft(trusted_profile_id)
        self.profile_authoring_service.update_default_omit_rules(
            draft_state.trusted_profile_draft_id,
            [{"phase_code": "20", "phase_name": "Labor"}],
            expected_draft_revision=draft_state.draft_revision,
        )
        parsed_record = self._make_material_record(vendor_name_normalized="Vendor A")

        with patch("services.review_workflow_service.parse_pdf", return_value=[parsed_record]):
            before_publish = self.processing_run_service.create_processing_run(self.source_document_path)

        published_detail = self.profile_authoring_service.publish_draft(draft_state.trusted_profile_draft_id)

        with patch("services.review_workflow_service.parse_pdf", return_value=[parsed_record]):
            after_publish = self.processing_run_service.create_processing_run(self.source_document_path)

        versions = self.lineage_store.list_trusted_profile_versions(trusted_profile_id)
        previous_version = self.lineage_store.get_trusted_profile_version(
            original_detail.current_published_version_id
        )

        self.assertEqual(len(versions), 2)
        self.assertNotEqual(
            published_detail.current_published_version_id,
            original_detail.current_published_version_id,
        )
        self.assertEqual(previous_version.version_number, 1)
        self.assertEqual(previous_version.bundle_payload["behavioral_bundle"]["review_rules"]["default_omit_rules"], [])
        self.assertNotEqual(
            before_publish.processing_run.trusted_profile_version_id,
            after_publish.processing_run.trusted_profile_version_id,
        )
        self.assertEqual(
            self.lineage_store.get_processing_run(before_publish.processing_run.processing_run_id).trusted_profile_version_id,
            original_detail.current_published_version_id,
        )

    def test_sqlite_import_preserves_ids_and_lineage_rows_in_postgres(self) -> None:
        sqlite_database_path = TEST_ROOT / "import-source.db"
        sqlite_store = SqliteLineageStore(sqlite_database_path)
        try:
            (
                sqlite_repository,
                sqlite_provisioning_service,
                sqlite_execution_adapter,
                sqlite_profile_authoring_service,
                sqlite_processing_run_service,
                sqlite_review_session_service,
            ) = self._build_service_stack(sqlite_store)

            with patch(
                "services.review_workflow_service.parse_pdf",
                return_value=[self._make_material_record(vendor_name_normalized="Vendor A")],
            ):
                processing_result = sqlite_processing_run_service.create_processing_run(self.source_document_path)

            updated_state = sqlite_review_session_service.apply_review_edits(
                processing_result.processing_run.processing_run_id,
                [
                    PendingRecordEdit(
                        record_key="record-0",
                        changed_fields={"vendor_name_normalized": "Vendor Imported"},
                    )
                ],
            )
            export_output = TEST_ROOT / "sqlite-export" / "revision-1.xlsx"
            export_output.parent.mkdir(parents=True, exist_ok=True)
            sqlite_review_session_service.export_session_revision(
                processing_result.processing_run.processing_run_id,
                session_revision=1,
                output_path=export_output,
            )
            source_profile = sqlite_repository.get_trusted_profile(
                "org-default",
                "trusted-profile:org-default:default",
            )
        finally:
            sqlite_store.close()

        import_schema = self._new_schema_name("import")
        imported_counts = import_sqlite_lineage_to_postgres(
            sqlite_database_path=sqlite_database_path,
            postgres_connection_string=self.postgres_pooled_url,
            migration_connection_string=self.postgres_admin_url,
            schema_name=import_schema,
        )
        imported_store = PostgresLineageStore(
            connection_string=self.postgres_pooled_url,
            migration_connection_string=self.postgres_admin_url,
            schema_name=import_schema,
            apply_migrations=False,
        )
        try:
            imported_profile = imported_store.get_trusted_profile(source_profile.trusted_profile_id)

            self.assertEqual(imported_counts["processing_runs"], 0)
            self.assertEqual(imported_counts["reviewed_record_edits"], 0)
            self.assertEqual(imported_counts["retained_export_artifacts"], 1)
            self.assertEqual(imported_profile.trusted_profile_id, source_profile.trusted_profile_id)
            with self.assertRaises(KeyError):
                imported_store.get_processing_run(processing_result.processing_run.processing_run_id)
            self.assertEqual(
                imported_store.list_export_artifacts(updated_state.review_session.review_session_id)[0].session_revision,
                1,
            )
        finally:
            imported_store.close()

    def test_postgres_org_aware_fetches_fail_closed_for_cross_org_reads(self) -> None:
        processing_result = self._create_processing_run()
        export_output = TEST_ROOT / "exports" / "cross-org.xlsx"
        export_output.parent.mkdir(parents=True, exist_ok=True)
        export_result = self.review_session_service.export_session_revision(
            processing_result.processing_run.processing_run_id,
            session_revision=0,
            output_path=export_output,
        )

        with self.assertRaises(KeyError):
            self.lineage_store.get_processing_run_for_organization(
                organization_id="org-alt",
                processing_run_id=processing_result.processing_run.processing_run_id,
            )
        with self.assertRaises(KeyError):
            self.lineage_store.get_trusted_profile_for_organization(
                organization_id="org-alt",
                trusted_profile_id=processing_result.trusted_profile.trusted_profile_id,
            )
        with self.assertRaises(KeyError):
            self.lineage_store.get_export_artifact_for_organization(
                organization_id="org-alt",
                export_artifact_id=export_result.export_artifact.export_artifact_id,
            )

    def _create_postgres_store(self) -> PostgresLineageStore:
        return PostgresLineageStore(
            connection_string=self.postgres_pooled_url,
            migration_connection_string=self.postgres_admin_url,
            schema_name=self._new_schema_name("main"),
        )

    def _build_service_stack(self, lineage_store):
        repository = TrustedProfileAuthoringRepository(
            lineage_store=lineage_store,
            now_provider=lambda: self.created_at,
        )
        trusted_profile_provisioning_service = TrustedProfileProvisioningService(
            lineage_store=lineage_store,
            repository=repository,
            profile_manager=self.profile_manager,
            now_provider=lambda: self.created_at,
        )
        profile_execution_compatibility_adapter = ProfileExecutionCompatibilityAdapter(
            lineage_store=lineage_store,
            profile_manager=self.profile_manager,
        )
        profile_authoring_service = ProfileAuthoringService(
            repository=repository,
            trusted_profile_provisioning_service=trusted_profile_provisioning_service,
            profile_execution_compatibility_adapter=profile_execution_compatibility_adapter,
            profile_manager=self.profile_manager,
            now_provider=lambda: self.created_at,
        )
        processing_run_service = ProcessingRunService(
            lineage_store=lineage_store,
            trusted_profile_provisioning_service=trusted_profile_provisioning_service,
            profile_execution_compatibility_adapter=profile_execution_compatibility_adapter,
            profile_authoring_service=profile_authoring_service,
            engine_version="engine-1",
            now_provider=lambda: self.created_at,
        )
        review_session_service = ReviewSessionService(
            lineage_store=lineage_store,
            profile_execution_compatibility_adapter=profile_execution_compatibility_adapter,
            now_provider=lambda: self.created_at,
        )
        return (
            repository,
            trusted_profile_provisioning_service,
            profile_execution_compatibility_adapter,
            profile_authoring_service,
            processing_run_service,
            review_session_service,
        )

    def _new_schema_name(self, suffix: str) -> str:
        schema_name = f"phase3_{suffix}_{uuid4().hex[:12]}"
        self.schema_names.append(schema_name)
        return schema_name

    def _drop_schema(self, schema_name: str) -> None:
        with psycopg.connect(self.postgres_admin_url) as connection:
            connection.execute(
                sql.SQL("DROP SCHEMA IF EXISTS {} CASCADE").format(sql.Identifier(schema_name))
            )

    def _create_processing_run(self):
        parsed_record = self._make_material_record(vendor_name_normalized="Vendor A")
        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[parsed_record],
        ):
            return self.processing_run_service.create_processing_run(self.source_document_path)

    def _write_profile_bundle(self) -> None:
        profile_dir = TEST_ROOT / "profiles" / "default"
        self._write_json(
            profile_dir / "profile.json",
            {
                "profile_name": "default",
                "display_name": "Default Profile",
                "description": "Default test profile",
                "version": "1.0",
                "template_filename": "recap_template.xlsx",
                "is_active": False,
            },
        )
        self._write_json(profile_dir / "labor_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "equipment_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "phase_mapping.json", {"50": "MATERIAL"})
        self._write_json(profile_dir / "vendor_normalization.json", {})
        self._write_json(
            profile_dir / "input_model.json",
            {"report_type": "vista_job_cost", "section_headers": {}},
        )
        self._write_json(
            profile_dir / "target_labor_classifications.json",
            {
                "slots": [{"slot_id": "labor_1", "label": "Default Journeyman", "active": True}],
                "classifications": ["Default Journeyman"],
            },
        )
        self._write_json(
            profile_dir / "target_equipment_classifications.json",
            {
                "slots": [{"slot_id": "equipment_1", "label": "Pick-up Truck", "active": True}],
                "classifications": ["Pick-up Truck"],
            },
        )
        self._write_json(profile_dir / "rates.json", {"labor_rates": {}, "equipment_rates": {}})
        self._write_json(profile_dir / "review_rules.json", {"default_omit_rules": []})
        self._write_json(
            profile_dir / "recap_template_map.json",
            {
                "worksheet_name": "Recap",
                "header_fields": {
                    "project": {"cell": "B6"},
                    "job_number": {"cell": "H6"},
                },
                "labor_rows": {
                    "Default Journeyman": {
                        "st_hours": "B14",
                        "ot_hours": "C14",
                        "dt_hours": "D14",
                        "st_rate": "E14",
                        "ot_rate": "F14",
                        "dt_rate": "G14",
                    }
                },
                "equipment_rows": {"Pick-up Truck": {"hours_qty": "B32", "rate": "D32"}},
                "materials_section": {
                    "start_row": 27,
                    "end_row": 41,
                    "columns": {"name": "G", "amount": "H"},
                },
                "subcontractors_section": {
                    "start_row": 46,
                    "end_row": 50,
                    "columns": {"name": "A", "amount": "C"},
                },
                "permits_fees_section": {
                    "start_row": 55,
                    "end_row": 56,
                    "columns": {"description": "A", "amount": "C"},
                },
                "police_detail_section": {
                    "start_row": 61,
                    "end_row": 62,
                    "columns": {"description": "A", "amount": "C"},
                },
                "sales_tax_area": {
                    "rate_label_cell": "G60",
                    "rate_input_cell": "H60",
                    "amount_label_cell": "G61",
                    "amount_formula_cell": "H61",
                    "material_total_cell": "H54",
                },
            },
        )
        self._create_template(profile_dir / "recap_template.xlsx")

    def _create_template(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Recap"
        for cell, value in {
            "A1": "Template V1",
            "A6": "Project",
            "G6": "Job Number",
            "H23": "=SUM(H12:H22)",
            "A25": "EQUIPMENT",
            "A26": "Category",
            "B26": "Hours / Qty",
            "D26": "Rate",
            "G25": "MATERIALS",
            "G26": "Vendor",
            "H26": "Amount",
            "E42": "=SUM(E27:E41)",
            "H42": "=SUM(H27:H41)",
            "C51": "=SUM(C46:C50)",
            "C57": "=SUM(C55:C56)",
            "C63": "=SUM(C61:C62)",
            "F58": 0,
        }.items():
            worksheet[cell] = value
        workbook.save(path)

    def _make_material_record(self, *, vendor_name_normalized: str) -> Record:
        return Record(
            record_type=MATERIAL,
            phase_code="50",
            raw_description="Material line",
            cost=100.0,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name="Vendor A",
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Material source",
            record_type_normalized=MATERIAL,
            recap_labor_classification=None,
            vendor_name_normalized=vendor_name_normalized,
        )

    def _write_json(self, path: Path, payload: object) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    unittest.main()
