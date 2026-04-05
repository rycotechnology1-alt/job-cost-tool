"""Semantic workbook snapshot helpers for parity acceptance."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def capture_workbook_snapshot(workbook_path: str | Path, expected_snapshot: dict[str, object]) -> dict[str, object]:
    """Capture the semantic workbook cells requested by one parity case."""
    workbook = load_workbook(filename=workbook_path, data_only=False)
    worksheet_name = str(expected_snapshot["worksheet_name"])
    worksheet = workbook[worksheet_name]

    captured_cells: dict[str, dict[str, Any]] = {}
    for cell_ref, cell_expectation in dict(expected_snapshot.get("cells", {})).items():
        cell = worksheet[cell_ref]
        captured_cells[cell_ref] = {
            "value": cell.value,
        }
        if isinstance(cell_expectation, dict) and "style_id" in cell_expectation:
            captured_cells[cell_ref]["style_id"] = cell.style_id
    return {
        "worksheet_name": worksheet_name,
        "cells": captured_cells,
    }


def capture_non_empty_workbook_snapshot(workbook_path: str | Path, worksheet_name: str = "Recap") -> dict[str, object]:
    """Capture a dense semantic snapshot of all non-empty cells on one worksheet."""
    workbook = load_workbook(filename=workbook_path, data_only=False)
    worksheet = workbook[worksheet_name]
    captured_cells: dict[str, dict[str, Any]] = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            captured_cells[cell.coordinate] = {
                "value": cell.value,
                "style_id": cell.style_id,
            }
    return {
        "worksheet_name": worksheet_name,
        "cells": captured_cells,
    }


def compare_workbook_snapshots(
    expected_snapshot: dict[str, object],
    actual_snapshot: dict[str, object],
    *,
    label: str,
) -> list[str]:
    """Return semantic workbook mismatch messages."""
    diffs: list[str] = []
    if expected_snapshot.get("worksheet_name") != actual_snapshot.get("worksheet_name"):
        diffs.append(
            f"{label}: worksheet mismatch expected {expected_snapshot.get('worksheet_name')!r} "
            f"but got {actual_snapshot.get('worksheet_name')!r}"
        )

    expected_cells = dict(expected_snapshot.get("cells", {}))
    actual_cells = dict(actual_snapshot.get("cells", {}))
    for cell_ref, expected_cell in expected_cells.items():
        actual_cell = actual_cells.get(cell_ref)
        if actual_cell is None:
            diffs.append(f"{label}: missing expected cell snapshot for {cell_ref}")
            continue
        if actual_cell.get("value") != expected_cell.get("value"):
            diffs.append(
                f"{label}: cell {cell_ref} value mismatch expected {expected_cell.get('value')!r} "
                f"but got {actual_cell.get('value')!r}"
            )
        if "style_id" in expected_cell and actual_cell.get("style_id") != expected_cell.get("style_id"):
            diffs.append(
                f"{label}: cell {cell_ref} style_id mismatch expected {expected_cell.get('style_id')!r} "
                f"but got {actual_cell.get('style_id')!r}"
            )
    return diffs
