"""Acceptance corpus loading and fixture materialization for phase-1 parity tests."""

from __future__ import annotations

import json
import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.models import Record


CORPUS_ROOT = Path("tests/parity_corpus")


@dataclass(frozen=True, slots=True)
class ParityCase:
    """One acceptance-corpus case for semantic desktop-versus-web comparison."""

    case_name: str
    case_dir: Path
    source_filename: str
    trusted_profile_name: str
    template_filename: str
    profile_source_dir: str | None
    description_filename: str | None
    reference_export_filename: str | None
    notes: str | None
    profile_bundle: dict[str, object] | None
    template_seed: dict[str, object] | None
    parsed_records: list[dict[str, object]] | None
    scripted_edit_batches: list[list[dict[str, object]]]
    target_export_revision: int | None
    expected_base_records: list[dict[str, object]] | None
    expected_base_blockers: list[str] | None
    expected_edited_records: list[dict[str, object]] | None
    expected_edited_blockers: list[str] | None
    expected_export_snapshot: dict[str, object] | None

    @property
    def source_path(self) -> Path:
        """Return the corpus source-document path."""
        return self.case_dir / self.source_filename

    @property
    def description_path(self) -> Path | None:
        """Return the optional corpus description path."""
        if not self.description_filename:
            return None
        return self.case_dir / self.description_filename

    @property
    def reference_export_path(self) -> Path | None:
        """Return the optional expected-export workbook path."""
        if not self.reference_export_filename:
            return None
        return self.case_dir / self.reference_export_filename

    def build_parsed_records(self) -> list[Record]:
        """Return parsed-record fixtures as Record objects for parser patching."""
        return [Record(**record_payload) for record_payload in (self.parsed_records or [])]


def load_parity_case(case_name: str) -> ParityCase:
    """Load one acceptance-corpus case by directory name."""
    case_dir = (CORPUS_ROOT / case_name).resolve()
    case_payload = json.loads((case_dir / "case.json").read_text(encoding="utf-8"))
    expected_payload = case_payload["expected"]
    return ParityCase(
        case_name=case_payload["case_name"],
        case_dir=case_dir,
        source_filename=case_payload["source_document"]["filename"],
        trusted_profile_name=case_payload["trusted_profile"]["profile_name"],
        template_filename=str(case_payload["trusted_profile"].get("template_filename") or "recap_template.xlsx"),
        profile_source_dir=case_payload["trusted_profile"].get("profile_source_dir"),
        description_filename=case_payload.get("description", {}).get("filename"),
        reference_export_filename=case_payload.get("reference_export", {}).get("filename"),
        notes=case_payload.get("notes"),
        profile_bundle=(
            dict(case_payload["trusted_profile"]["bundle"])
            if "bundle" in case_payload["trusted_profile"]
            else None
        ),
        template_seed=(
            dict(case_payload["trusted_profile"]["template_seed"])
            if "template_seed" in case_payload["trusted_profile"]
            else None
        ),
        parsed_records=list(case_payload["parsed_records"]) if "parsed_records" in case_payload else None,
        scripted_edit_batches=list(case_payload["scripted_edit_batches"]),
        target_export_revision=case_payload.get("target_export_revision"),
        expected_base_records=list(expected_payload["base_review"]["records"]) if "base_review" in expected_payload else None,
        expected_base_blockers=list(expected_payload["base_review"]["blocking_issues"]) if "base_review" in expected_payload else None,
        expected_edited_records=list(expected_payload["edited_review"]["records"]) if "edited_review" in expected_payload else None,
        expected_edited_blockers=list(expected_payload["edited_review"]["blocking_issues"]) if "edited_review" in expected_payload else None,
        expected_export_snapshot=dict(expected_payload["export_workbook"]) if "export_workbook" in expected_payload else None,
    )


def materialize_case_runtime(case: ParityCase, runtime_root: Path) -> tuple[Path, Path, Path]:
    """Create one isolated runtime workspace for a parity case."""
    if runtime_root.exists():
        shutil.rmtree(runtime_root)
    profile_dir = runtime_root / "profiles" / case.trusted_profile_name
    legacy_config_dir = runtime_root / "legacy_config"
    legacy_config_dir.mkdir(parents=True, exist_ok=True)

    if case.profile_source_dir:
        source_profile_dir = (case.case_dir / case.profile_source_dir).resolve()
        shutil.copytree(source_profile_dir, profile_dir)
    else:
        profile_dir.mkdir(parents=True, exist_ok=True)
        if case.profile_bundle is None or case.template_seed is None:
            raise ValueError(f"Parity case '{case.case_name}' does not define a trusted profile fixture.")
        for file_name, payload in case.profile_bundle.items():
            target_path = profile_dir / file_name
            target_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        _create_template_from_seed(profile_dir / str(case.template_seed["filename"]), case.template_seed)

    (legacy_config_dir / "phase_catalog.json").write_text('{"phases":[]}', encoding="utf-8")

    source_path = runtime_root / case.source_filename
    source_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(case.source_path, source_path)
    return profile_dir, legacy_config_dir, source_path


def _create_template_from_seed(path: Path, template_seed: dict[str, object]) -> None:
    """Build one deterministic XLSX template from a JSON seed."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = str(template_seed["worksheet_name"])

    cells = dict(template_seed.get("cells", {}))
    for cell_ref, cell_spec in cells.items():
        if isinstance(cell_spec, dict):
            worksheet[cell_ref] = cell_spec.get("value")
            _apply_cell_style(worksheet[cell_ref], dict(cell_spec.get("style", {})))
        else:
            worksheet[cell_ref] = cell_spec

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _apply_cell_style(cell, style_payload: dict[str, object]) -> None:
    """Apply a minimal deterministic style set from JSON seed data."""
    fill_color = str(style_payload.get("fill") or "").strip()
    if fill_color:
        cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

    if style_payload.get("bold"):
        cell.font = Font(bold=True)

    number_format = str(style_payload.get("number_format") or "").strip()
    if number_format:
        cell.number_format = number_format
