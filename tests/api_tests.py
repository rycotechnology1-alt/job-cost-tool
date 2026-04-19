"""API tests for the narrow phase-1 FastAPI backend slice."""

from __future__ import annotations

import base64
import hashlib
import hmac
import io
import json
import os
import shutil
import unittest
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from api import create_app
from core.config import ConfigLoader, ProfileManager
from core.models import MATERIAL, Record
from core.models.lineage import TrustedProfile
from infrastructure.persistence import SqliteLineageStore
from infrastructure.storage import VercelBlobRuntimeStorage
from services.profile_authoring_errors import ProfileAuthoringPersistenceConflictError
from services.request_context import RequestContext
from tests.runtime_storage_test_helpers import FakeBlobObjectClient


TEST_ROOT = Path("tests/_api_tmp")


class Phase1ApiTests(unittest.TestCase):
    """Verify the minimal FastAPI slice delegates to the accepted lineage services cleanly."""

    def setUp(self) -> None:
        ConfigLoader.clear_runtime_caches()
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        (TEST_ROOT / "profiles" / "default").mkdir(parents=True, exist_ok=True)
        (TEST_ROOT / "profiles" / "alternate").mkdir(parents=True, exist_ok=True)
        (TEST_ROOT / "legacy_config").mkdir(parents=True, exist_ok=True)
        self.settings_path = TEST_ROOT / "app_settings.json"
        self.created_at = datetime(2026, 4, 5, 12, 0, tzinfo=timezone.utc)
        self.current_time = self.created_at

        self._write_profile_bundle()
        self._write_profile_bundle(
            profile_name="alternate",
            display_name="Alternate Profile",
            description="Alternate API test profile",
            template_filename="alternate_template.xlsx",
            labor_classifications=["ALT Journeyman"],
            equipment_classifications=["ALT Truck"],
        )
        self._write_json(self.settings_path, {"active_profile": "default"})
        self._write_json(TEST_ROOT / "legacy_config" / "phase_catalog.json", {"phases": []})

        self.profile_manager = ProfileManager(
            profiles_root=TEST_ROOT / "profiles",
            legacy_config_root=TEST_ROOT / "legacy_config",
        )
        self.lineage_store = SqliteLineageStore()
        self.app = create_app(
            lineage_store=self.lineage_store,
            database_provider="sqlite",
            profile_manager=self.profile_manager,
            storage_provider="local",
            upload_root=TEST_ROOT / "runtime" / "uploads",
            export_root=TEST_ROOT / "runtime" / "exports",
            upload_retention_hours=24,
            engine_version="engine-1",
            now_provider=lambda: self.current_time,
        )
        self.client = TestClient(self.app)
        provisioning_service = self.app.state.runtime.trusted_profile_service._trusted_profile_provisioning_service
        organization = provisioning_service._ensure_request_organization(None)
        provisioning_service._bootstrap_new_filesystem_profile(
            organization=organization,
            profile_name="alternate",
            metadata=self.profile_manager.get_profile_metadata("alternate"),
        )

    def tearDown(self) -> None:
        self.client.close()
        self.lineage_store.close()
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        ConfigLoader.clear_runtime_caches()

    def test_source_upload_and_processing_run_creation(self) -> None:
        upload_response = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_payload = upload_response.json()

        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[self._make_material_record(vendor_name_normalized="Vendor A")],
        ):
            first_run_response = self.client.post(
                "/api/runs",
                json={
                    "upload_id": upload_payload["upload_id"],
                    "trusted_profile_name": "default",
                },
            )
            second_run_response = self.client.post(
                "/api/runs",
                json={
                    "upload_id": upload_payload["upload_id"],
                    "trusted_profile_name": "default",
                },
            )

        self.assertEqual(first_run_response.status_code, 201)
        self.assertEqual(second_run_response.status_code, 201)
        first_payload = first_run_response.json()
        second_payload = second_run_response.json()

        first_run_detail = self.client.get(f"/api/runs/{first_payload['processing_run_id']}")
        second_run_detail = self.client.get(f"/api/runs/{second_payload['processing_run_id']}")

        self.assertEqual(first_run_detail.status_code, 200)
        self.assertEqual(second_run_detail.status_code, 200)
        self.assertEqual(first_payload["record_count"], 1)
        self.assertEqual(first_payload["aggregate_blockers"], [])
        self.assertEqual(first_payload["source_document_filename"], "report.pdf")
        self.assertTrue(first_payload["historical_export_status"]["is_reproducible"])
        self.assertEqual(first_run_detail.json()["run_records"][0]["record_key"], "record-0")
        self.assertEqual(first_run_detail.json()["source_document_filename"], "report.pdf")
        self.assertEqual(second_run_detail.json()["run_records"][0]["record_key"], "record-0")
        self.assertNotEqual(first_payload["processing_run_id"], second_payload["processing_run_id"])

    def test_trusted_profile_listing_returns_read_only_phase1_selection_metadata(self) -> None:
        response = self.client.get("/api/trusted-profiles")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual([profile["profile_name"] for profile in payload], ["alternate", "default"])
        self.assertEqual(payload[0]["display_name"], "Alternate Profile")
        self.assertEqual(payload[0]["source_kind"], "filesystem_bootstrap")
        self.assertFalse(payload[0]["has_open_draft"])
        self.assertFalse(payload[0]["is_active_profile"])
        self.assertIsNone(payload[0]["archived_at"])
        self.assertEqual(payload[1]["trusted_profile_id"], "trusted-profile:org-default:default")
        self.assertEqual(payload[1]["source_kind"], "seeded")
        self.assertEqual(payload[1]["current_published_version_number"], 1)
        self.assertTrue(payload[1]["is_active_profile"])
        self.assertIsNone(payload[1]["archived_at"])
        self.assertEqual(payload[1]["template_filename"], "recap_template.xlsx")

    def test_request_context_provider_can_scope_listing_to_a_different_organization_without_auth(self) -> None:
        self._set_request_context_org("org-alt")

        response = self.client.get("/api/trusted-profiles")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([profile["profile_name"] for profile in payload], ["default"])
        self.assertEqual(payload[0]["trusted_profile_id"], "trusted-profile:org-alt:default")

    def test_bearer_auth_requires_token_for_hosted_api(self) -> None:
        hosted_client = self._create_hosted_client()
        try:
            response = hosted_client.get("/api/trusted-profiles")
        finally:
            hosted_client.close()

        self.assertEqual(response.status_code, 401)

    def test_bearer_auth_provisions_org_user_and_scoped_default_profile(self) -> None:
        self._write_json(self.settings_path, {"active_profile": "alternate"})
        hosted_client = self._create_hosted_client()
        try:
            response = hosted_client.get(
                "/api/trusted-profiles",
                headers=self._auth_headers(
                    organization_id="org-acme",
                    organization_slug="acme",
                    organization_name="Acme Organization",
                    user_id="user-acme-1",
                    email="acme.user@example.com",
                    display_name="Acme User",
                    role="member",
                ),
            )
        finally:
            hosted_client.close()

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([profile["profile_name"] for profile in payload], ["default"])
        self.assertEqual(payload[0]["trusted_profile_id"], "trusted-profile:org-acme:default")
        self.assertTrue(payload[0]["is_active_profile"])

        user_row = self.lineage_store._connection.execute(
            "SELECT organization_id, email, auth_subject FROM users WHERE user_id = ?",
            ("user-acme-1",),
        ).fetchone()
        organization_row = self.lineage_store._connection.execute(
            "SELECT slug, display_name, default_trusted_profile_id FROM organizations WHERE organization_id = ?",
            ("org-acme",),
        ).fetchone()

        self.assertEqual(user_row["organization_id"], "org-acme")
        self.assertEqual(user_row["email"], "acme.user@example.com")
        self.assertEqual(user_row["auth_subject"], "auth-user-acme-1")
        self.assertEqual(organization_row["slug"], "acme")
        self.assertEqual(organization_row["display_name"], "Acme Organization")
        self.assertEqual(
            organization_row["default_trusted_profile_id"],
            "trusted-profile:org-acme:default",
        )

    def test_bearer_auth_with_local_sentinel_values_still_uses_hosted_org_profile_resolution(self) -> None:
        self._write_json(self.settings_path, {"active_profile": "alternate"})
        hosted_client = self._create_hosted_client()
        try:
            response = hosted_client.get(
                "/api/trusted-profiles",
                headers=self._auth_headers(
                    organization_id="org-acme",
                    organization_slug="acme",
                    organization_name="Acme Organization",
                    user_id="dev-local-user",
                    email="dev.local@example.com",
                    display_name="Dev Local User",
                    role="developer",
                ),
            )
        finally:
            hosted_client.close()

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([profile["profile_name"] for profile in payload], ["default"])
        self.assertEqual(payload[0]["trusted_profile_id"], "trusted-profile:org-acme:default")

    def test_hosted_profile_creation_without_seed_uses_org_default_not_local_active_profile(self) -> None:
        self._write_json(self.settings_path, {"active_profile": "alternate"})
        hosted_client = self._create_hosted_client()
        headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="user-acme-1",
            email="acme.user@example.com",
            display_name="Acme User",
            role="member",
        )
        try:
            bootstrap_response = hosted_client.get("/api/trusted-profiles", headers=headers)
            create_response = hosted_client.post(
                "/api/profiles",
                headers=headers,
                json={
                    "profile_name": "field-team",
                    "display_name": "Field Team",
                    "description": "Created from hosted default profile",
                },
            )
        finally:
            hosted_client.close()

        self.assertEqual(bootstrap_response.status_code, 200)
        self.assertEqual(create_response.status_code, 201)
        self.assertEqual(
            create_response.json()["current_published_version"]["template_filename"],
            "recap_template.xlsx",
        )

    def test_bearer_auth_same_org_access_succeeds_and_cross_org_access_fails_closed(self) -> None:
        hosted_client = self._create_hosted_client()
        acme_headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="user-acme-1",
            email="acme.user@example.com",
            display_name="Acme User",
            role="member",
        )
        beta_headers = self._auth_headers(
            organization_id="org-beta",
            organization_slug="beta",
            organization_name="Beta Organization",
            user_id="user-beta-1",
            email="beta.user@example.com",
            display_name="Beta User",
            role="member",
        )
        try:
            processing_run_id = self._create_processing_run_via_api(
                client=hosted_client,
                headers=acme_headers,
            )
            same_org_response = hosted_client.get(f"/api/runs/{processing_run_id}", headers=acme_headers)
            cross_org_response = hosted_client.get(f"/api/runs/{processing_run_id}", headers=beta_headers)
        finally:
            hosted_client.close()

        self.assertEqual(same_org_response.status_code, 200)
        self.assertEqual(cross_org_response.status_code, 404)

    def test_processing_run_routes_fail_closed_for_cross_org_reads(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        self._set_request_context_org("org-alt")

        run_response = self.client.get(f"/api/runs/{processing_run_id}")
        review_response = self.client.get(f"/api/runs/{processing_run_id}/review-session")

        self.assertEqual(run_response.status_code, 404)
        self.assertEqual(review_response.status_code, 404)
        self.assertIn(processing_run_id, run_response.json()["detail"])

    def test_hosted_processing_run_with_local_sentinel_values_still_persists_authenticated_user_id(self) -> None:
        hosted_client = self._create_hosted_client()
        headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="dev-local-user",
            email="dev.local@example.com",
            display_name="Dev Local User",
            role="developer",
        )
        try:
            processing_run_id = self._create_processing_run_via_api(
                client=hosted_client,
                headers=headers,
            )
        finally:
            hosted_client.close()

        persisted_run = self.lineage_store.get_processing_run(processing_run_id)
        self.assertEqual(persisted_run.organization_id, "org-acme")
        self.assertEqual(persisted_run.created_by_user_id, "dev-local-user")

    def test_profile_draft_routes_fail_closed_for_cross_org_reads(self) -> None:
        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()
        draft_id = draft_payload["trusted_profile_draft_id"]
        self.assertEqual(draft_payload["draft_revision"], 1)

        self._set_request_context_org("org-alt")

        detail_response = self.client.get(f"/api/profile-drafts/{draft_id}")
        publish_response = self.client.post(
            f"/api/profile-drafts/{draft_id}/publish",
            json={"expected_draft_revision": draft_payload["draft_revision"]},
        )

        self.assertEqual(detail_response.status_code, 404)
        self.assertEqual(publish_response.status_code, 404)
        self.assertIn(draft_id, detail_response.json()["detail"])

    def test_create_second_profile_lists_opens_saves_and_publishes_independently(self) -> None:
        create_response = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "field-team",
                "display_name": "Field Team",
                "description": "Second trusted profile",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )
        self.assertEqual(create_response.status_code, 201)
        created_payload = create_response.json()

        listing_response = self.client.get("/api/trusted-profiles")
        self.assertEqual(listing_response.status_code, 200)
        listing_payload = listing_response.json()
        self.assertEqual([profile["profile_name"] for profile in listing_payload], ["alternate", "default", "field-team"])
        self.assertEqual(listing_payload[2]["source_kind"], "published_clone")
        self.assertFalse(listing_payload[2]["has_open_draft"])

        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:field-team/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()
        draft_id = draft_payload["trusted_profile_draft_id"]
        self.assertEqual(draft_payload["draft_revision"], 1)

        listing_with_draft = self.client.get("/api/trusted-profiles")
        self.assertTrue(listing_with_draft.json()[2]["has_open_draft"])

        save_response = self.client.patch(
            f"/api/profile-drafts/{draft_id}/default-omit",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "default_omit_rules": [
                    {"phase_code": "50", "phase_name": "Other Job Cost"},
                ]
            },
        )
        self.assertEqual(save_response.status_code, 200)

        publish_response = self.client.post(
            f"/api/profile-drafts/{draft_id}/publish",
            json={"expected_draft_revision": save_response.json()["draft_revision"]},
        )
        self.assertEqual(publish_response.status_code, 200)
        published_payload = publish_response.json()

        default_detail = self.client.get("/api/profiles/trusted-profile:org-default:default")
        created_detail = self.client.get("/api/profiles/trusted-profile:org-default:field-team")

        self.assertEqual(created_payload["current_published_version"]["version_number"], 1)
        self.assertEqual(published_payload["current_published_version"]["version_number"], 2)
        self.assertEqual(
            published_payload["current_published_version"]["trusted_profile_version_id"],
            created_detail.json()["current_published_version"]["trusted_profile_version_id"],
        )
        self.assertEqual(default_detail.status_code, 200)
        self.assertEqual(default_detail.json()["current_published_version"]["version_number"], 1)
        self.assertIsNone(published_payload["open_draft_id"])
        self.assertEqual(published_payload["profile_name"], "field-team")

    def test_profile_draft_patch_returns_conflict_for_stale_revision(self) -> None:
        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()

        first_response = self.client.patch(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/export-settings",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "export_settings": {
                    "labor_minimum_hours": {
                        "enabled": True,
                        "threshold_hours": "2",
                        "minimum_hours": "4",
                    }
                },
            },
        )
        stale_response = self.client.patch(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/export-settings",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "export_settings": {
                    "labor_minimum_hours": {
                        "enabled": False,
                        "threshold_hours": "",
                        "minimum_hours": "",
                    }
                },
            },
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(stale_response.status_code, 409)
        self.assertEqual(
            stale_response.json()["detail"]["error_code"],
            "profile_authoring_persistence_conflict",
        )

    def test_profile_draft_publish_returns_conflict_for_stale_revision(self) -> None:
        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()

        save_response = self.client.patch(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/default-omit",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "default_omit_rules": [
                    {"phase_code": "20", "phase_name": "Labor"},
                ],
            },
        )
        self.assertEqual(save_response.status_code, 200)

        stale_publish_response = self.client.post(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/publish",
            json={"expected_draft_revision": draft_payload["draft_revision"]},
        )

        self.assertEqual(stale_publish_response.status_code, 409)
        self.assertEqual(
            stale_publish_response.json()["detail"]["error_code"],
            "profile_authoring_persistence_conflict",
        )

    def test_profile_draft_publish_retry_returns_conflict_after_first_publish_succeeds(self) -> None:
        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()

        save_response = self.client.patch(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/default-omit",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "default_omit_rules": [
                    {"phase_code": "20", "phase_name": "Labor"},
                ],
            },
        )
        self.assertEqual(save_response.status_code, 200)
        current_revision = save_response.json()["draft_revision"]

        first_publish_response = self.client.post(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/publish",
            json={"expected_draft_revision": current_revision},
        )
        retry_publish_response = self.client.post(
            f"/api/profile-drafts/{draft_payload['trusted_profile_draft_id']}/publish",
            json={"expected_draft_revision": current_revision},
        )

        self.assertEqual(first_publish_response.status_code, 200)
        self.assertEqual(retry_publish_response.status_code, 409)
        self.assertEqual(
            retry_publish_response.json()["detail"]["error_code"],
            "profile_authoring_persistence_conflict",
        )

    def test_create_second_profile_rejects_duplicate_key_and_duplicate_active_display_name(self) -> None:
        first_create = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "field-team",
                "display_name": "Field Team",
                "description": "Second trusted profile",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )
        self.assertEqual(first_create.status_code, 201)

        duplicate_key_response = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "FIELD-TEAM",
                "display_name": "Field Team 2",
                "description": "Duplicate key",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )
        duplicate_display_response = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "field-team-2",
                "display_name": "Field Team",
                "description": "Duplicate display name",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )

        self.assertEqual(duplicate_key_response.status_code, 409)
        self.assertIn("already exists", duplicate_key_response.json()["detail"]["message"])
        self.assertIn("profile_name", duplicate_key_response.json()["detail"]["field_errors"])
        self.assertEqual(duplicate_display_response.status_code, 409)
        self.assertIn("already in use", duplicate_display_response.json()["detail"]["message"])
        self.assertIn("display_name", duplicate_display_response.json()["detail"]["field_errors"])

    def test_archive_user_created_profile_hides_it_from_active_listing_without_deleting_history(self) -> None:
        create_response = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "field-team",
                "display_name": "Field Team",
                "description": "Second trusted profile",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )
        self.assertEqual(create_response.status_code, 201)

        archive_response = self.client.post("/api/profiles/trusted-profile:org-default:field-team/archive")
        self.assertEqual(archive_response.status_code, 204)

        listing_response = self.client.get("/api/trusted-profiles")
        detail_response = self.client.get("/api/profiles/trusted-profile:org-default:field-team")
        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:field-team/draft")

        self.assertEqual([profile["profile_name"] for profile in listing_response.json()], ["alternate", "default"])
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()["current_published_version"]["version_number"], 1)
        self.assertEqual(draft_response.status_code, 400)
        self.assertIn("archived", draft_response.json()["detail"])

    def test_trusted_profile_listing_can_include_archived_profiles_and_unarchive_restores_active_listing(self) -> None:
        create_response = self.client.post(
            "/api/profiles",
            json={
                "profile_name": "field-team",
                "display_name": "Field Team",
                "description": "Second trusted profile",
                "seed_trusted_profile_id": "trusted-profile:org-default:default",
            },
        )
        self.assertEqual(create_response.status_code, 201)

        archive_response = self.client.post("/api/profiles/trusted-profile:org-default:field-team/archive")
        self.assertEqual(archive_response.status_code, 204)

        active_listing = self.client.get("/api/trusted-profiles")
        all_listing = self.client.get("/api/trusted-profiles?include_archived=true")

        self.assertEqual([profile["profile_name"] for profile in active_listing.json()], ["alternate", "default"])
        archived_profile = next(
            profile for profile in all_listing.json() if profile["profile_name"] == "field-team"
        )
        self.assertTrue(archived_profile["archived_at"])

        unarchive_response = self.client.post("/api/profiles/trusted-profile:org-default:field-team/unarchive")
        self.assertEqual(unarchive_response.status_code, 204)

        restored_listing = self.client.get("/api/trusted-profiles")
        self.assertEqual(
            [profile["profile_name"] for profile in restored_listing.json()],
            ["alternate", "default", "field-team"],
        )

    def test_review_session_open_and_append_edit_batch_do_not_mutate_base_run_records(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        session_response = self.client.get(f"/api/runs/{processing_run_id}/review-session")
        self.assertEqual(session_response.status_code, 200)
        self.assertEqual(session_response.json()["current_revision"], 0)
        self.assertEqual(session_response.json()["session_revision"], 0)
        self.assertEqual(session_response.json()["labor_classification_options"], ["103 Journeyman"])
        self.assertEqual(session_response.json()["equipment_classification_options"], ["Pick-up Truck"])
        self.assertTrue(session_response.json()["historical_export_status"]["is_reproducible"])

        edit_response = self.client.post(
            f"/api/runs/{processing_run_id}/review-session/edits",
            json={
                "edits": [
                    {
                        "record_key": "record-0",
                        "changed_fields": {"vendor_name_normalized": "Vendor Edited"},
                    }
                ]
            },
        )
        self.assertEqual(edit_response.status_code, 200)
        edit_payload = edit_response.json()

        run_detail_response = self.client.get(f"/api/runs/{processing_run_id}")
        self.assertEqual(run_detail_response.status_code, 200)
        run_payload = run_detail_response.json()

        self.assertEqual(edit_payload["current_revision"], 1)
        self.assertEqual(edit_payload["session_revision"], 1)
        self.assertEqual(edit_payload["records"][0]["vendor_name_normalized"], "Vendor Edited")
        self.assertEqual(
            run_payload["run_records"][0]["canonical_record"]["vendor_name_normalized"],
            "Vendor A",
        )

    def test_review_session_open_uses_processing_run_profile_snapshot_for_option_sets(self) -> None:
        processing_run_id = self._create_processing_run_via_api(trusted_profile_name="alternate")

        session_response = self.client.get(f"/api/runs/{processing_run_id}/review-session")

        self.assertEqual(session_response.status_code, 200)
        self.assertEqual(session_response.json()["labor_classification_options"], ["ALT Journeyman"])
        self.assertEqual(session_response.json()["equipment_classification_options"], ["ALT Truck"])

    def test_review_edit_rejects_invalid_classification_submission(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        edit_response = self.client.post(
            f"/api/runs/{processing_run_id}/review-session/edits",
            json={
                "edits": [
                    {
                        "record_key": "record-0",
                        "changed_fields": {"recap_labor_classification": "Not Allowed"},
                    }
                ]
            },
        )
        session_response = self.client.get(f"/api/runs/{processing_run_id}/review-session")

        self.assertEqual(edit_response.status_code, 400)
        self.assertIn("not allowed for this review", edit_response.json()["detail"])
        self.assertEqual(session_response.status_code, 200)
        self.assertEqual(session_response.json()["current_revision"], 0)

    def test_hosted_review_edit_returns_conflict_for_stale_expected_current_revision(self) -> None:
        hosted_client = self._create_hosted_client()
        headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="user-acme-1",
            email="acme.user@example.com",
            display_name="Acme User",
            role="member",
        )
        try:
            processing_run_id = self._create_processing_run_via_api(
                client=hosted_client,
                headers=headers,
            )
            session_response = hosted_client.get(
                f"/api/runs/{processing_run_id}/review-session",
                headers=headers,
            )
            self.assertEqual(session_response.status_code, 200)

            first_edit_response = hosted_client.post(
                f"/api/runs/{processing_run_id}/review-session/edits",
                headers=headers,
                json={
                    "expected_current_revision": session_response.json()["current_revision"],
                    "edits": [
                        {
                            "record_key": "record-0",
                            "changed_fields": {"vendor_name_normalized": "Vendor Rev 1"},
                        }
                    ],
                },
            )
            stale_edit_response = hosted_client.post(
                f"/api/runs/{processing_run_id}/review-session/edits",
                headers=headers,
                json={
                    "expected_current_revision": session_response.json()["current_revision"],
                    "edits": [
                        {
                            "record_key": "record-0",
                            "changed_fields": {"vendor_name_normalized": "Vendor Rev 2"},
                        }
                    ],
                },
            )
        finally:
            hosted_client.close()

        self.assertEqual(first_edit_response.status_code, 200)
        self.assertEqual(stale_edit_response.status_code, 409)
        self.assertEqual(
            stale_edit_response.json()["detail"]["error_code"],
            "review_session_persistence_conflict",
        )

    def test_hosted_review_edit_requires_expected_current_revision(self) -> None:
        hosted_client = self._create_hosted_client()
        headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="user-acme-1",
            email="acme.user@example.com",
            display_name="Acme User",
            role="member",
        )
        try:
            processing_run_id = self._create_processing_run_via_api(
                client=hosted_client,
                headers=headers,
            )
            response = hosted_client.post(
                f"/api/runs/{processing_run_id}/review-session/edits",
                headers=headers,
                json={
                    "edits": [
                        {
                            "record_key": "record-0",
                            "changed_fields": {"vendor_name_normalized": "Vendor Rev 1"},
                        }
                    ],
                },
            )
        finally:
            hosted_client.close()

        self.assertEqual(response.status_code, 400)
        self.assertIn("expected_current_revision", response.json()["detail"])

    def test_hosted_review_edit_with_local_sentinel_values_still_requires_expected_current_revision(self) -> None:
        hosted_client = self._create_hosted_client()
        headers = self._auth_headers(
            organization_id="org-acme",
            organization_slug="acme",
            organization_name="Acme Organization",
            user_id="dev-local-user",
            email="dev.local@example.com",
            display_name="Dev Local User",
            role="developer",
        )
        try:
            processing_run_id = self._create_processing_run_via_api(
                client=hosted_client,
                headers=headers,
            )
            response = hosted_client.post(
                f"/api/runs/{processing_run_id}/review-session/edits",
                headers=headers,
                json={
                    "edits": [
                        {
                            "record_key": "record-0",
                            "changed_fields": {"vendor_name_normalized": "Vendor Rev 1"},
                        }
                    ],
                },
            )
        finally:
            hosted_client.close()

        self.assertEqual(response.status_code, 400)
        self.assertIn("expected_current_revision", response.json()["detail"])

    def test_expired_upload_returns_clear_reupload_message_when_creating_run(self) -> None:
        upload_response = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_id = upload_response.json()["upload_id"]

        self.current_time += timedelta(hours=25)
        run_response = self.client.post(
            "/api/runs",
            json={
                "upload_id": upload_id,
                "trusted_profile_name": "default",
            },
        )

        self.assertEqual(run_response.status_code, 410)
        self.assertIn("expired from temporary storage", run_response.json()["detail"])
        self.assertIn("upload the PDF again", run_response.json()["detail"])

    def test_existing_run_remains_fetchable_after_original_upload_is_cleaned_up(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        self.current_time += timedelta(hours=25)
        second_upload = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report-2.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(second_upload.status_code, 201)

        run_detail_response = self.client.get(f"/api/runs/{processing_run_id}")
        review_session_response = self.client.get(f"/api/runs/{processing_run_id}/review-session")

        self.assertEqual(run_detail_response.status_code, 200)
        self.assertEqual(run_detail_response.json()["source_document_filename"], "report.pdf")
        self.assertEqual(review_session_response.status_code, 200)
        self.assertEqual(review_session_response.json()["current_revision"], 0)

    def test_export_creation_and_download_bind_to_one_exact_session_revision(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        first_edit = self.client.post(
            f"/api/runs/{processing_run_id}/review-session/edits",
            json={
                "edits": [
                    {
                        "record_key": "record-0",
                        "changed_fields": {"vendor_name_normalized": "Vendor Rev 1"},
                    }
                ]
            },
        )
        second_edit = self.client.post(
            f"/api/runs/{processing_run_id}/review-session/edits",
            json={
                "edits": [
                    {
                        "record_key": "record-0",
                        "changed_fields": {"vendor_name_normalized": "Vendor Rev 2"},
                    }
                ]
            },
        )
        self.assertEqual(first_edit.status_code, 200)
        self.assertEqual(second_edit.status_code, 200)

        export_response = self.client.post(
            f"/api/runs/{processing_run_id}/exports",
            json={"session_revision": 1},
        )
        self.assertEqual(export_response.status_code, 201)
        export_payload = export_response.json()

        download_response = self.client.get(export_payload["download_url"])
        self.assertEqual(download_response.status_code, 200)

        workbook = load_workbook(io.BytesIO(download_response.content))
        worksheet = workbook["Recap"]

        self.assertEqual(export_payload["session_revision"], 1)
        self.assertTrue(export_payload["template_artifact_id"])
        self.assertEqual(worksheet["G27"].value, "Vendor Rev 1")
        self.assertNotEqual(worksheet["G27"].value, "Vendor Rev 2")
        self.assertIn('filename="report-recap-rev-1.xlsx"', download_response.headers["content-disposition"])

    def test_export_artifact_download_fails_closed_for_cross_org_reads(self) -> None:
        processing_run_id = self._create_processing_run_via_api()

        export_response = self.client.post(
            f"/api/runs/{processing_run_id}/exports",
            json={"session_revision": 0},
        )
        self.assertEqual(export_response.status_code, 201)
        download_url = export_response.json()["download_url"]

        self._set_request_context_org("org-alt")

        download_response = self.client.get(download_url)

        self.assertEqual(download_response.status_code, 404)
        self.assertIn(export_response.json()["export_artifact_id"], download_response.json()["detail"])

    def test_legacy_runs_are_reported_non_reproducible_and_exact_export_fails_closed(self) -> None:
        processing_run_id = self._create_processing_run_via_api()
        run_detail = self.client.get(f"/api/runs/{processing_run_id}")
        self.assertEqual(run_detail.status_code, 200)
        snapshot_id = run_detail.json()["profile_snapshot_id"]

        self.lineage_store._connection.execute(
            """
            UPDATE profile_snapshots
            SET template_artifact_id = NULL,
                template_file_hash = NULL
            WHERE profile_snapshot_id = ?
            """,
            (snapshot_id,),
        )
        self.lineage_store._connection.commit()

        refreshed_run_detail = self.client.get(f"/api/runs/{processing_run_id}")
        review_session = self.client.get(f"/api/runs/{processing_run_id}/review-session")
        export_response = self.client.post(
            f"/api/runs/{processing_run_id}/exports",
            json={"session_revision": 0},
        )

        self.assertEqual(refreshed_run_detail.status_code, 200)
        self.assertEqual(
            refreshed_run_detail.json()["historical_export_status"]["status_code"],
            "legacy_non_reproducible",
        )
        self.assertEqual(review_session.status_code, 200)
        self.assertFalse(review_session.json()["historical_export_status"]["is_reproducible"])
        self.assertEqual(export_response.status_code, 409)
        self.assertIn("predates template-artifact capture", export_response.json()["detail"])

    def test_profile_authoring_endpoints_open_edit_and_publish_draft(self) -> None:
        self._write_json(
            TEST_ROOT / "profiles" / "default" / "equipment_mapping.json",
            {
                "raw_mappings": {"pickup truck": "Pick-up Truck"},
                "saved_mappings": [
                    {
                        "raw_description": "pickup truck",
                        "target_category": "Pick-up Truck",
                    }
                ],
            },
        )

        detail_response = self.client.get("/api/profiles/trusted-profile:org-default:default")
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(detail_response.json()["current_published_version"]["version_number"], 1)
        self.assertIn("phase_mapping", detail_response.json()["deferred_domains"])

        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_payload = draft_response.json()
        draft_id = draft_payload["trusted_profile_draft_id"]
        self.assertEqual(
            draft_payload["equipment_mappings"][0],
            {
                "raw_description": "PICKUP TRUCK",
                "raw_pattern": "PICKUP TRUCK",
                "target_category": "Pick-up Truck",
                "is_observed": False,
                "is_required_for_recent_processing": False,
                "prediction_target": None,
                "prediction_confidence_label": None,
            },
        )

        labor_mapping_response = self.client.patch(
            f"/api/profile-drafts/{draft_id}/labor-mappings",
            json={
                "expected_draft_revision": draft_payload["draft_revision"],
                "labor_mappings": [
                    {
                        "raw_value": "103/J",
                        "target_classification": "103 Journeyman",
                        "notes": "Mapped in web authoring",
                    }
                ]
            },
        )
        self.assertEqual(labor_mapping_response.status_code, 200)
        self.assertEqual(
            labor_mapping_response.json()["labor_mappings"][0]["notes"],
            "Mapped in web authoring",
        )
        self.assertEqual(labor_mapping_response.json()["draft_revision"], draft_payload["draft_revision"] + 1)

        equipment_mapping_response = self.client.patch(
            f"/api/profile-drafts/{draft_id}/equipment-mappings",
            json={
                "expected_draft_revision": labor_mapping_response.json()["draft_revision"],
                "equipment_mappings": [
                    {
                        "raw_description": "pickup truck",
                        "raw_pattern": "PICKUP TRUCK",
                        "target_category": "Pick-up Truck",
                    }
                ]
            },
        )
        self.assertEqual(equipment_mapping_response.status_code, 200)
        self.assertEqual(
            equipment_mapping_response.json()["equipment_mappings"][0],
            {
                "raw_description": "PICKUP TRUCK",
                "raw_pattern": "PICKUP TRUCK",
                "target_category": "Pick-up Truck",
                "is_observed": False,
                "is_required_for_recent_processing": False,
                "prediction_target": None,
                "prediction_confidence_label": None,
            },
        )

        publish_response = self.client.post(
            f"/api/profile-drafts/{draft_id}/publish",
            json={"expected_draft_revision": equipment_mapping_response.json()["draft_revision"]},
        )
        self.assertEqual(publish_response.status_code, 200)
        publish_payload = publish_response.json()
        self.assertEqual(publish_payload["current_published_version"]["version_number"], 2)
        self.assertIsNone(publish_payload["open_draft_id"])

        draft_detail_response = self.client.get(f"/api/profile-drafts/{draft_id}")
        self.assertEqual(draft_detail_response.status_code, 404)

    def test_profile_draft_delete_discards_open_draft_without_changing_published_version(self) -> None:
        detail_before = self.client.get("/api/profiles/trusted-profile:org-default:default")
        self.assertEqual(detail_before.status_code, 200)
        version_before = detail_before.json()["current_published_version"]["trusted_profile_version_id"]

        draft_response = self.client.post("/api/profiles/trusted-profile:org-default:default/draft")
        self.assertEqual(draft_response.status_code, 201)
        draft_id = draft_response.json()["trusted_profile_draft_id"]

        discard_response = self.client.delete(f"/api/profile-drafts/{draft_id}")
        self.assertEqual(discard_response.status_code, 204)

        detail_after = self.client.get("/api/profiles/trusted-profile:org-default:default")
        missing_draft = self.client.get(f"/api/profile-drafts/{draft_id}")

        self.assertEqual(detail_after.status_code, 200)
        self.assertEqual(
            detail_after.json()["current_published_version"]["trusted_profile_version_id"],
            version_before,
        )
        self.assertIsNone(detail_after.json()["open_draft_id"])
        self.assertEqual(missing_draft.status_code, 404)

    def test_profile_detail_repairs_default_profile_missing_current_published_version(self) -> None:
        organization = self.lineage_store.ensure_organization(
            organization_id="org-default",
            slug="default-org",
            display_name="Default Organization",
            created_at=self.created_at,
            is_seeded=True,
        )
        self.lineage_store.get_or_create_trusted_profile(
            TrustedProfile(
                trusted_profile_id="trusted-profile:org-default:default",
                organization_id=organization.organization_id,
                profile_name="default",
                display_name="Default Profile",
                source_kind="seeded",
                bundle_ref=str(TEST_ROOT / "profiles" / "default"),
                description="Default API test profile",
                version_label="1.0",
                created_at=self.created_at,
            )
        )

        response = self.client.get("/api/profiles/trusted-profile:org-default:default")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["current_published_version"]["version_number"], 1)
        repaired_profile = self.lineage_store.get_trusted_profile("trusted-profile:org-default:default")
        persisted_versions = self.lineage_store.list_trusted_profile_versions(
            repaired_profile.trusted_profile_id
        )
        self.assertEqual(len(persisted_versions), 1)
        self.assertEqual(
            repaired_profile.current_published_version_id,
            persisted_versions[0].trusted_profile_version_id,
        )

    def test_processing_run_observation_capture_exposes_observed_blank_row_in_draft_state(self) -> None:
        upload_response = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_payload = upload_response.json()

        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[self._make_unmapped_labor_record(raw_description="Labor line", union_code="104", labor_class_raw="EO")],
        ):
            run_response = self.client.post(
                "/api/runs",
                json={
                    "upload_id": upload_payload["upload_id"],
                    "trusted_profile_name": "default",
                },
            )

        self.assertEqual(run_response.status_code, 201)
        profile_detail = self.client.get("/api/profiles/trusted-profile:org-default:default")
        self.assertEqual(profile_detail.status_code, 200)
        draft_id = profile_detail.json()["open_draft_id"]
        self.assertTrue(draft_id)

        draft_response = self.client.get(f"/api/profile-drafts/{draft_id}")
        self.assertEqual(draft_response.status_code, 200)
        self.assertIn(
            {
                "raw_value": "104/EO",
                "target_classification": "",
                "notes": "",
                "is_observed": True,
                "is_required_for_recent_processing": True,
            },
            draft_response.json()["labor_mappings"],
        )

    def test_processing_run_returns_created_and_records_all_observations_when_one_draft_merge_is_stale(self) -> None:
        upload_response = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_payload = upload_response.json()

        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[
                self._make_unmapped_labor_record(raw_description="Labor line", union_code="104", labor_class_raw="EO"),
                self._make_unmapped_equipment_record(raw_description="627/2025 crane truck"),
            ],
        ):
            with patch.object(
                self.client.app.state.runtime.profile_authoring_service,
                "_save_validated_bundle",
                side_effect=ProfileAuthoringPersistenceConflictError("stale draft"),
            ) as save_mock:
                run_response = self.client.post(
                    "/api/runs",
                    json={
                        "upload_id": upload_payload["upload_id"],
                        "trusted_profile_name": "default",
                    },
                )

        self.assertEqual(run_response.status_code, 201)
        run_payload = run_response.json()
        persisted_runs = self.lineage_store.list_processing_runs()
        observations = self.lineage_store.list_trusted_profile_observations("trusted-profile:org-default:default")

        self.assertEqual(len(persisted_runs), 1)
        self.assertEqual(persisted_runs[0].processing_run_id, run_payload["processing_run_id"])
        self.assertEqual(
            {observation.canonical_raw_key for observation in observations},
            {"104/EO", "CRANE TRUCK"},
        )
        self.assertEqual(save_mock.call_count, 2)

    def test_processing_run_observation_capture_exposes_equipment_prediction_metadata_in_draft_state(self) -> None:
        self._write_json(
            TEST_ROOT / "profiles" / "default" / "equipment_mapping.json",
            {
                "raw_mappings": {"pickup truck": "Pick-up Truck"},
                "saved_mappings": [
                    {
                        "raw_description": "pickup truck",
                        "target_category": "Pick-up Truck",
                    }
                ],
            },
        )
        upload_response = self.client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_payload = upload_response.json()

        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[self._make_unmapped_equipment_record(raw_description="pickup")],
        ):
            run_response = self.client.post(
                "/api/runs",
                json={
                    "upload_id": upload_payload["upload_id"],
                    "trusted_profile_name": "default",
                },
            )

        self.assertEqual(run_response.status_code, 201)
        profile_detail = self.client.get("/api/profiles/trusted-profile:org-default:default")
        self.assertEqual(profile_detail.status_code, 200)
        draft_id = profile_detail.json()["open_draft_id"]
        self.assertTrue(draft_id)

        draft_response = self.client.get(f"/api/profile-drafts/{draft_id}")
        self.assertEqual(draft_response.status_code, 200)
        self.assertIn(
            {
                "raw_description": "PICKUP",
                "raw_pattern": "PICKUP",
                "target_category": "",
                "is_observed": True,
                "is_required_for_recent_processing": True,
                "prediction_target": "Pick-up Truck",
                "prediction_confidence_label": "High confidence",
            },
            draft_response.json()["equipment_mappings"],
        )

    def test_profile_sync_endpoint_is_not_registered(self) -> None:
        version_id = "trusted-profile-version:org-default:default:v1"

        response = self.client.post(f"/api/profile-versions/{version_id}/desktop-sync-export")

        self.assertEqual(response.status_code, 404)

    def test_shared_blob_storage_supports_cross_instance_upload_process_export_and_download(self) -> None:
        shared_blob_client = FakeBlobObjectClient()
        instance_one = self._create_multi_instance_client(
            runtime_root=TEST_ROOT / "blob-instance-one",
            blob_client=shared_blob_client,
        )
        instance_two = self._create_multi_instance_client(
            runtime_root=TEST_ROOT / "blob-instance-two",
            blob_client=shared_blob_client,
        )
        try:
            upload_response = instance_one.post(
                "/api/source-documents/uploads",
                files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
            )
            self.assertEqual(upload_response.status_code, 201)

            with patch(
                "services.review_workflow_service.parse_pdf",
                return_value=[self._make_material_record(vendor_name_normalized="Vendor A")],
            ):
                run_response = instance_two.post(
                    "/api/runs",
                    json={
                        "upload_id": upload_response.json()["upload_id"],
                        "trusted_profile_name": "default",
                    },
                )
            self.assertEqual(run_response.status_code, 201)
            processing_run_id = run_response.json()["processing_run_id"]

            export_response = instance_one.post(
                f"/api/runs/{processing_run_id}/exports",
                json={"session_revision": 0},
            )
            self.assertEqual(export_response.status_code, 201)
            export_download = instance_two.get(export_response.json()["download_url"])
            self.assertEqual(export_download.status_code, 200)
            self.assertTrue(export_download.content)

        finally:
            instance_one.close()
            instance_two.close()

    def _create_processing_run_via_api(
        self,
        *,
        trusted_profile_name: str = "default",
        client: TestClient | None = None,
        headers: dict[str, str] | None = None,
    ) -> str:
        active_client = client or self.client
        request_headers = headers or {}
        upload_response = active_client.post(
            "/api/source-documents/uploads",
            files={"file": ("report.pdf", b"sample pdf bytes", "application/pdf")},
            headers=request_headers,
        )
        self.assertEqual(upload_response.status_code, 201)
        upload_payload = upload_response.json()

        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[self._make_material_record(vendor_name_normalized="Vendor A")],
        ):
            run_response = active_client.post(
                "/api/runs",
                json={
                    "upload_id": upload_payload["upload_id"],
                    "trusted_profile_name": trusted_profile_name,
                },
                headers=request_headers,
            )

        self.assertEqual(run_response.status_code, 201)
        return run_response.json()["processing_run_id"]

    def _create_hosted_client(self, *, auth_secret: str = "test-auth-secret") -> TestClient:
        with patch.dict(
            os.environ,
            {
                "JOB_COST_API_AUTH_MODE": "bearer",
                "JOB_COST_API_AUTH_SECRET": auth_secret,
            },
            clear=False,
        ):
            return TestClient(
                create_app(
                    lineage_store=self.lineage_store,
                    database_provider="sqlite",
                    profile_manager=self.profile_manager,
                    storage_provider="local",
                    upload_root=TEST_ROOT / "runtime" / "uploads",
                    export_root=TEST_ROOT / "runtime" / "exports",
                    upload_retention_hours=24,
                    engine_version="engine-1",
                    now_provider=lambda: self.current_time,
                )
            )

    def _create_multi_instance_client(
        self,
        *,
        runtime_root: Path,
        blob_client: FakeBlobObjectClient,
    ) -> TestClient:
        file_store = VercelBlobRuntimeStorage(
            blob_client=blob_client,
            upload_root=runtime_root / "uploads",
            export_root=runtime_root / "exports",
            upload_retention_hours=24,
            now_provider=lambda: self.current_time,
        )
        return TestClient(
            create_app(
                lineage_store=self.lineage_store,
                profile_manager=self.profile_manager,
                file_store=file_store,
                upload_root=runtime_root / "uploads",
                export_root=runtime_root / "exports",
                upload_retention_hours=24,
                engine_version="engine-1",
                now_provider=lambda: self.current_time,
            )
        )

    def _auth_headers(
        self,
        *,
        organization_id: str,
        organization_slug: str,
        organization_name: str,
        user_id: str,
        email: str,
        display_name: str,
        role: str,
        auth_secret: str = "test-auth-secret",
    ) -> dict[str, str]:
        payload = {
            "sub": f"auth-{user_id}",
            "user_id": user_id,
            "email": email,
            "display_name": display_name,
            "organization_id": organization_id,
            "organization_slug": organization_slug,
            "organization_name": organization_name,
            "role": role,
        }
        payload_json = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
        encoded_payload = base64.urlsafe_b64encode(payload_json).rstrip(b"=").decode("ascii")
        signature = hmac.new(
            auth_secret.encode("utf-8"),
            encoded_payload.encode("ascii"),
            hashlib.sha256,
        ).digest()
        encoded_signature = base64.urlsafe_b64encode(signature).rstrip(b"=").decode("ascii")
        return {"Authorization": f"Bearer jobcostv1.{encoded_payload}.{encoded_signature}"}

    def _set_request_context_org(self, organization_id: str) -> None:
        self.client.app.state.request_context_provider = (
            lambda request: RequestContext(
                organization_id=organization_id,
                user_id="dev-local-user",
                role="developer",
            )
        )

    def _write_profile_bundle(
        self,
        *,
        profile_name: str = "default",
        display_name: str = "Default Profile",
        description: str = "Default API test profile",
        template_filename: str = "recap_template.xlsx",
        labor_classifications: list[str] | None = None,
        equipment_classifications: list[str] | None = None,
    ) -> None:
        profile_dir = TEST_ROOT / "profiles" / profile_name
        resolved_labor_classifications = labor_classifications or ["103 Journeyman"]
        resolved_equipment_classifications = equipment_classifications or ["Pick-up Truck"]
        self._write_json(
            profile_dir / "profile.json",
            {
                "profile_name": profile_name,
                "display_name": display_name,
                "description": description,
                "version": "1.0",
                "template_filename": template_filename,
                "is_active": False,
            },
        )
        self._write_json(profile_dir / "labor_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "equipment_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "phase_mapping.json", {"50": "MATERIAL"})
        self._write_json(profile_dir / "vendor_normalization.json", {})
        self._write_json(
            profile_dir / "input_model.json",
            {"report_type": "vista_job_cost", "section_headers": {}},
        )
        self._write_json(
            profile_dir / "target_labor_classifications.json",
            {
                "slots": [
                    {"slot_id": f"labor_{index + 1}", "label": label, "active": True}
                    for index, label in enumerate(resolved_labor_classifications)
                ],
                "classifications": resolved_labor_classifications,
            },
        )
        self._write_json(
            profile_dir / "target_equipment_classifications.json",
            {
                "slots": [
                    {"slot_id": f"equipment_{index + 1}", "label": label, "active": True}
                    for index, label in enumerate(resolved_equipment_classifications)
                ],
                "classifications": resolved_equipment_classifications,
            },
        )
        self._write_json(profile_dir / "rates.json", {"labor_rates": {}, "equipment_rates": {}})
        self._write_json(profile_dir / "review_rules.json", {"default_omit_rules": []})
        self._write_json(
            profile_dir / "recap_template_map.json",
            {
                "worksheet_name": "Recap",
                "header_fields": {
                    "project": {"cell": "B6"},
                    "job_number": {"cell": "H6"},
                },
                "labor_rows": {
                    resolved_labor_classifications[0]: {
                        "st_hours": "B14",
                        "ot_hours": "C14",
                        "dt_hours": "D14",
                        "st_rate": "E14",
                        "ot_rate": "F14",
                        "dt_rate": "G14",
                    }
                },
                "equipment_rows": {resolved_equipment_classifications[0]: {"hours_qty": "B32", "rate": "D32"}},
                "materials_section": {
                    "start_row": 27,
                    "end_row": 41,
                    "columns": {"name": "G", "amount": "H"},
                },
                "subcontractors_section": {
                    "start_row": 46,
                    "end_row": 50,
                    "columns": {"name": "A", "amount": "C"},
                },
                "permits_fees_section": {
                    "start_row": 55,
                    "end_row": 56,
                    "columns": {"description": "A", "amount": "C"},
                },
                "police_detail_section": {
                    "start_row": 61,
                    "end_row": 62,
                    "columns": {"description": "A", "amount": "C"},
                },
                "sales_tax_area": {
                    "rate_label_cell": "G60",
                    "rate_input_cell": "H60",
                    "amount_label_cell": "G61",
                    "amount_formula_cell": "H61",
                    "material_total_cell": "H54",
                },
            },
        )
        self._create_template(profile_dir / template_filename)

    def _create_template(self, path: Path) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Recap"
        for cell, value in {
            "A6": "Project",
            "G6": "Job Number",
            "H23": "=SUM(H12:H22)",
            "A25": "EQUIPMENT",
            "A26": "Category",
            "B26": "Hours / Qty",
            "D26": "Rate",
            "G25": "MATERIALS",
            "G26": "Vendor",
            "H26": "Amount",
            "E42": "=SUM(E27:E41)",
            "H42": "=SUM(H27:H41)",
            "C51": "=SUM(C46:C50)",
            "C57": "=SUM(C55:C56)",
            "C63": "=SUM(C61:C62)",
            "F58": 0,
        }.items():
            worksheet[cell] = value
        workbook.save(path)

    def _make_material_record(self, *, vendor_name_normalized: str) -> Record:
        return Record(
            record_type=MATERIAL,
            phase_code="50",
            raw_description="Material line",
            cost=100.0,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name="Vendor A",
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Material source",
            record_type_normalized=MATERIAL,
            recap_labor_classification=None,
            vendor_name_normalized=vendor_name_normalized,
        )

    def _make_unmapped_labor_record(
        self,
        *,
        raw_description: str,
        union_code: str,
        labor_class_raw: str,
    ) -> Record:
        return Record(
            record_type="labor",
            phase_code="20",
            raw_description=raw_description,
            cost=100.0,
            hours=8.0,
            hour_type="ST",
            union_code=union_code,
            labor_class_raw=labor_class_raw,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=["Labor raw value is not mapped."],
            source_page=1,
            source_line_text="Labor source",
            record_type_normalized="labor",
            recap_labor_classification=None,
        )

    def _make_unmapped_equipment_record(self, *, raw_description: str) -> Record:
        return Record(
            record_type="equipment",
            phase_code="20",
            raw_description=raw_description,
            cost=100.0,
            hours=8.0,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description=raw_description,
            equipment_category=None,
            confidence=0.9,
            warnings=["Equipment raw value is not mapped."],
            source_page=1,
            source_line_text="Equipment source",
            record_type_normalized="equipment",
            equipment_mapping_key=raw_description,
        )

    def _write_json(self, path: Path, payload: object) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    unittest.main()
