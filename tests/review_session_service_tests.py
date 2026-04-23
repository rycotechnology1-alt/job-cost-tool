"""Service-level tests for review-session overlays and exact-revision export lineage."""

from __future__ import annotations

import json
import shutil
import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from core.config import ConfigLoader, ProfileManager
from core.models import EQUIPMENT, LABOR, MATERIAL, PendingRecordEdit, Record
from core.models.lineage import User
from infrastructure.persistence.sqlite_lineage_store import SqliteLineageStore
from infrastructure.storage import StoredArtifact
from infrastructure.storage import LocalRuntimeFileStore
from services.profile_execution_compatibility_adapter import ProfileExecutionCompatibilityAdapter
from services.profile_authoring_errors import ProfileAuthoringPersistenceConflictError
from services.processing_run_service import ProcessingRunService
from services.profile_authoring_service import ProfileAuthoringService
from services.request_context import RequestContext
from services.review_session_service import HistoricalExportUnavailableError, ReviewSessionService
from services.trusted_profile_authoring_repository import TrustedProfileAuthoringRepository
from services.trusted_profile_provisioning_service import TrustedProfileProvisioningService


TEST_ROOT = Path("tests/_review_session_tmp")


class ReviewSessionServiceTests(unittest.TestCase):
    """Verify append-only review overlays and exact-revision export generation."""

    def setUp(self) -> None:
        ConfigLoader.clear_runtime_caches()
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        (TEST_ROOT / "profiles" / "default").mkdir(parents=True, exist_ok=True)
        (TEST_ROOT / "legacy_config").mkdir(parents=True, exist_ok=True)
        self.settings_path = TEST_ROOT / "app_settings.json"
        self.source_document_path = TEST_ROOT / "sample_report.pdf"
        self.source_document_path.write_bytes(b"sample pdf bytes")
        self.created_at = datetime(2026, 4, 5, 12, 0, tzinfo=timezone.utc)

        self._write_profile_bundle()
        self._write_json(self.settings_path, {"active_profile": "default"})
        self._write_json(TEST_ROOT / "legacy_config" / "phase_catalog.json", {"phases": []})

        self.profile_manager = ProfileManager(
            profiles_root=TEST_ROOT / "profiles",
            legacy_config_root=TEST_ROOT / "legacy_config",
        )
        self.lineage_store = SqliteLineageStore()
        self.repository = TrustedProfileAuthoringRepository(
            lineage_store=self.lineage_store,
            now_provider=lambda: self.created_at,
        )
        self.trusted_profile_provisioning_service = TrustedProfileProvisioningService(
            lineage_store=self.lineage_store,
            repository=self.repository,
            profile_manager=self.profile_manager,
            now_provider=lambda: self.created_at,
        )
        self.profile_execution_compatibility_adapter = ProfileExecutionCompatibilityAdapter(
            lineage_store=self.lineage_store,
            profile_manager=self.profile_manager,
        )
        self.profile_authoring_service = ProfileAuthoringService(
            repository=self.repository,
            trusted_profile_provisioning_service=self.trusted_profile_provisioning_service,
            profile_execution_compatibility_adapter=self.profile_execution_compatibility_adapter,
            profile_manager=self.profile_manager,
            now_provider=lambda: self.created_at,
        )
        self.processing_run_service = ProcessingRunService(
            lineage_store=self.lineage_store,
            trusted_profile_provisioning_service=self.trusted_profile_provisioning_service,
            profile_execution_compatibility_adapter=self.profile_execution_compatibility_adapter,
            profile_authoring_service=self.profile_authoring_service,
            engine_version="engine-1",
            now_provider=lambda: self.created_at,
        )
        self.review_session_service = ReviewSessionService(
            lineage_store=self.lineage_store,
            profile_execution_compatibility_adapter=self.profile_execution_compatibility_adapter,
            now_provider=lambda: self.created_at,
        )

    def tearDown(self) -> None:
        self.lineage_store.close()
        shutil.rmtree(TEST_ROOT, ignore_errors=True)
        ConfigLoader.clear_runtime_caches()

    def test_review_edits_are_persisted_as_overlays_without_mutating_run_records(self) -> None:
        processing_result = self._create_processing_run()

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )

        persisted_run_records = self.lineage_store.list_run_records(processing_result.processing_run.processing_run_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(updated_state.review_session.review_session_id)

        self.assertEqual(updated_state.review_session.current_revision, 1)
        self.assertEqual(updated_state.session_revision, 1)
        self.assertEqual(updated_state.records[0].vendor_name_normalized, "Vendor B")
        self.assertEqual(persisted_run_records[0].canonical_record["vendor_name_normalized"], "Vendor A")
        self.assertEqual(persisted_edits[0].changed_fields, {"vendor_name_normalized": "Vendor B"})

    def test_hosted_review_edits_reject_stale_expected_current_revision(self) -> None:
        processing_result = self._create_processing_run()
        self.lineage_store.ensure_user(
            User(
                user_id="user-1",
                organization_id="org-default",
                email="user1@example.com",
                display_name="User One",
                auth_subject="auth-user-1",
                is_active=True,
                created_at=self.created_at,
            )
        )
        request_context = RequestContext(
            organization_id="org-default",
            user_id="user-1",
            role="member",
        )

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
            expected_current_revision=0,
            request_context=request_context,
        )

        with self.assertRaises(ProfileAuthoringPersistenceConflictError) as exc_info:
            self.review_session_service.apply_review_edits(
                processing_result.processing_run.processing_run_id,
                [
                    PendingRecordEdit(
                        record_key="record-0",
                        changed_fields={"vendor_name_normalized": "Vendor C"},
                    )
                ],
                expected_current_revision=0,
                request_context=request_context,
            )

        self.assertEqual(updated_state.review_session.current_revision, 1)
        self.assertEqual(exc_info.exception.error_code, "review_session_persistence_conflict")
        self.assertIn("expected_current_revision", exc_info.exception.field_errors)

    def test_reopening_review_session_resumes_latest_revision_for_the_run(self) -> None:
        processing_result = self._create_processing_run()
        self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )

        reopened_state = self.review_session_service.open_review_session(
            processing_result.processing_run.processing_run_id,
        )

        self.assertEqual(reopened_state.review_session.current_revision, 1)
        self.assertEqual(reopened_state.session_revision, 1)
        self.assertEqual(reopened_state.records[0].vendor_name_normalized, "Vendor B")

    def test_reopen_original_processed_previews_revision_zero_without_mutating_latest_revision(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )

        preview_state = self.review_session_service.reopen_review_session(
            processing_run_id,
            mode="original_processed",
        )
        latest_state = self.review_session_service.open_review_session(processing_run_id)

        self.assertEqual(preview_state.effective_source_mode, "original_processed")
        self.assertEqual(preview_state.session_revision, 0)
        self.assertEqual(preview_state.review_session.current_revision, 1)
        self.assertEqual(preview_state.records[0].vendor_name_normalized, "Vendor A")
        self.assertEqual(latest_state.review_session.current_revision, 1)
        self.assertEqual(latest_state.records[0].vendor_name_normalized, "Vendor B")

    def test_continue_from_original_appends_reset_overlay_and_makes_original_latest_working_state(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )

        reset_state = self.review_session_service.reopen_review_session(
            processing_run_id,
            mode="original_processed",
            continue_from_original=True,
        )
        reopened_state = self.review_session_service.open_review_session(processing_run_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(reset_state.review_session.review_session_id)

        self.assertEqual(reset_state.effective_source_mode, "latest_reviewed")
        self.assertEqual(reset_state.review_session.current_revision, 2)
        self.assertEqual(reset_state.session_revision, 2)
        self.assertEqual(reset_state.records[0].vendor_name_normalized, "Vendor A")
        self.assertEqual(reopened_state.review_session.current_revision, 2)
        self.assertEqual(reopened_state.records[0].vendor_name_normalized, "Vendor A")
        self.assertEqual([edit.session_revision for edit in persisted_edits], [1, 2])
        self.assertEqual(persisted_edits[-1].changed_fields["vendor_name_normalized"], "Vendor A")

    def test_exports_before_and_after_reset_remain_discoverable_in_history(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        review_session_id = self.review_session_service.open_review_session(processing_run_id).review_session.review_session_id

        first_output = TEST_ROOT / "exports" / "before-reset.xlsx"
        first_output.parent.mkdir(parents=True, exist_ok=True)
        self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=0,
            output_path=first_output,
        )
        self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor B"},
                )
            ],
        )
        self.review_session_service.reopen_review_session(
            processing_run_id,
            mode="original_processed",
            continue_from_original=True,
        )
        second_output = TEST_ROOT / "exports" / "after-reset.xlsx"
        self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=2,
            output_path=second_output,
        )

        persisted_artifacts = self.lineage_store.list_export_artifacts(review_session_id)

        self.assertEqual([artifact.session_revision for artifact in persisted_artifacts], [0, 2])

    def test_review_session_state_exposes_snapshot_classification_options(self) -> None:
        processing_result = self._create_processing_run()

        state = self.review_session_service.open_review_session(
            processing_result.processing_run.processing_run_id,
        )

        self.assertEqual(
            state.labor_classification_options,
            ["103 General FM", "103 Foreman", "103 Journeyman"],
        )
        self.assertEqual(state.equipment_classification_options, ["Pick-up Truck"])

    def test_manual_equipment_category_fix_clears_review_warnings_without_mutating_run_record(self) -> None:
        processing_result = self._create_processing_run_with_record(
            self._make_unmapped_equipment_record(
                warnings=["PR equipment detail line was recognized but equipment description was not parsed cleanly."],
            ),
        )
        processing_run_id = processing_result.processing_run.processing_run_id

        initial_state = self.review_session_service.open_review_session(processing_run_id)
        updated_state = self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"equipment_category": "Pick-up Truck"},
                )
            ],
        )

        persisted_run_records = self.lineage_store.list_run_records(processing_run_id)

        self.assertIn("Equipment recap category is missing.", "\n".join(initial_state.records[0].warnings))
        self.assertEqual(updated_state.records[0].equipment_category, "Pick-up Truck")
        self.assertEqual(updated_state.records[0].warnings, [])
        self.assertEqual(updated_state.blocking_issues, [])
        self.assertEqual(persisted_run_records[0].canonical_record["equipment_category"], None)
        self.assertIn(
            "Equipment description did not match a configured target equipment category.",
            persisted_run_records[0].canonical_record["warnings"],
        )
        self.assertIn(
            "PR equipment detail line was recognized but equipment description was not parsed cleanly.",
            persisted_run_records[0].canonical_record["warnings"],
        )

    def test_manual_labor_fix_clears_mapping_warning_and_medium_confidence_review_flag(self) -> None:
        processing_result = self._create_processing_run_with_record(
            self._make_unmapped_labor_record(
                warnings=["PR labor detail line was recognized but labor class was not parsed cleanly."],
            ),
        )

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"recap_labor_classification": "103 Journeyman"},
                )
            ],
        )

        self.assertEqual(updated_state.records[0].recap_labor_classification, "103 Journeyman")
        self.assertEqual(updated_state.records[0].warnings, [])
        self.assertEqual(updated_state.blocking_issues, [])

    def test_reprocessed_labor_mapping_resolution_hides_parser_warning_in_effective_review_state(self) -> None:
        parsed_record = self._make_unmapped_labor_record(
            warnings=["PR labor detail line was recognized but labor class was not parsed cleanly."],
        )
        first_result = self._create_processing_run_with_record(parsed_record)
        first_state = self.review_session_service.open_review_session(first_result.processing_run.processing_run_id)

        trusted_profile = self.trusted_profile_provisioning_service.resolve_current_published_profile()
        draft_state = self.profile_authoring_service.create_or_open_draft(trusted_profile.trusted_profile.trusted_profile_id)
        updated_draft = self.profile_authoring_service.update_labor_mappings(
            draft_state.trusted_profile_draft_id,
            [
                {
                    "raw_value": "103/ZZ",
                    "target_classification": "103 Journeyman",
                    "notes": "Mapped after observation",
                }
            ],
            expected_draft_revision=draft_state.draft_revision,
        )
        self.profile_authoring_service.publish_draft(
            updated_draft.trusted_profile_draft_id,
            expected_draft_revision=updated_draft.draft_revision,
        )

        second_result = self._create_processing_run_with_record(parsed_record)
        second_state = self.review_session_service.open_review_session(second_result.processing_run.processing_run_id)
        persisted_run_records = self.lineage_store.list_run_records(second_result.processing_run.processing_run_id)

        self.assertIn(
            "PR labor detail line was recognized but labor class was not parsed cleanly.",
            first_state.records[0].warnings,
        )
        self.assertIn(
            "Labor raw value '103/ZZ' is not mapped to a target recap labor classification.",
            first_state.records[0].warnings,
        )
        self.assertEqual(second_state.records[0].recap_labor_classification, "103 Journeyman")
        self.assertEqual(second_state.records[0].warnings, [])
        self.assertEqual(second_state.blocking_issues, [])
        self.assertIn(
            "PR labor detail line was recognized but labor class was not parsed cleanly.",
            persisted_run_records[0].canonical_record["warnings"],
        )

    def test_manual_vendor_fix_clears_vendor_resolution_warning(self) -> None:
        processing_result = self._create_processing_run_with_record(
            self._make_missing_vendor_material_record(),
        )

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor Reviewed"},
                )
            ],
        )

        self.assertEqual(updated_state.records[0].vendor_name_normalized, "Vendor Reviewed")
        self.assertEqual(updated_state.records[0].warnings, [])
        self.assertEqual(updated_state.blocking_issues, [])

    def test_manual_omit_clears_effective_review_warnings_and_blockers(self) -> None:
        processing_result = self._create_processing_run_with_record(
            self._make_unmapped_equipment_record(),
        )

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"is_omitted": True},
                )
            ],
        )

        self.assertTrue(updated_state.records[0].is_omitted)
        self.assertEqual(updated_state.records[0].warnings, [])
        self.assertEqual(updated_state.blocking_issues, [])

    def test_manual_fix_preserves_unrelated_ambiguity_warning(self) -> None:
        processing_result = self._create_processing_run_with_record(
            self._make_unmapped_equipment_record(
                warnings=["PR detail line family is ambiguous and should be reviewed."],
            ),
        )

        updated_state = self.review_session_service.apply_review_edits(
            processing_result.processing_run.processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"equipment_category": "Pick-up Truck"},
                )
            ],
        )

        self.assertIn("PR detail line family is ambiguous and should be reviewed.", updated_state.records[0].warnings)
        self.assertNotIn(
            "Equipment description did not match a configured target equipment category.",
            updated_state.records[0].warnings,
        )
        self.assertNotIn("Medium-confidence record should be reviewed before export.", updated_state.records[0].warnings)
        self.assertEqual(
            updated_state.blocking_issues,
            ["Record on page 1 (unknown phase, equipment): Record still contains unresolved parsing or normalization ambiguity."],
        )

    def test_export_uses_one_exact_session_revision_even_after_later_edits_exist(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id

        self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor Rev 1"},
                )
            ],
        )
        self.review_session_service.apply_review_edits(
            processing_run_id,
            [
                PendingRecordEdit(
                    record_key="record-0",
                    changed_fields={"vendor_name_normalized": "Vendor Rev 2"},
                )
            ],
        )

        revision_one_output = TEST_ROOT / "exports" / "revision-1.xlsx"
        revision_one_output.parent.mkdir(parents=True, exist_ok=True)
        export_result = self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=1,
            output_path=revision_one_output,
        )

        worksheet = load_workbook(revision_one_output)["Recap"]
        persisted_artifacts = self.lineage_store.list_export_artifacts(
            export_result.review_session_state.review_session.review_session_id,
        )

        self.assertEqual(export_result.export_artifact.session_revision, 1)
        self.assertEqual(
            export_result.export_artifact.template_artifact_id,
            processing_result.profile_snapshot.template_artifact_id,
        )
        self.assertEqual(export_result.review_session_state.review_session.current_revision, 2)
        self.assertEqual(export_result.review_session_state.session_revision, 1)
        self.assertEqual(worksheet["G27"].value, "Vendor Rev 1")
        self.assertEqual(persisted_artifacts[0].session_revision, 1)
        self.assertEqual(persisted_artifacts[0].processing_run_id, processing_run_id)

    def test_historical_export_uses_persisted_template_artifact_even_if_on_disk_workbook_changes(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id

        original_snapshot = self.lineage_store.get_profile_snapshot(
            processing_result.processing_run.profile_snapshot_id,
        )
        original_template_artifact = self.lineage_store.get_template_artifact(
            original_snapshot.template_artifact_id,
        )

        self._create_template(
            TEST_ROOT / "profiles" / "default" / "recap_template.xlsx",
            marker_value="Template V2 Mutated",
        )

        output_path = TEST_ROOT / "exports" / "historical.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        export_result = self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=0,
            output_path=output_path,
        )

        worksheet = load_workbook(output_path)["Recap"]

        self.assertEqual(original_template_artifact.original_filename, "recap_template.xlsx")
        self.assertEqual(
            export_result.export_artifact.template_artifact_id,
            original_snapshot.template_artifact_id,
        )
        self.assertEqual(worksheet["A1"].value, "Template V1")
        self.assertNotEqual(worksheet["A1"].value, "Template V2 Mutated")

    def test_historical_export_preserves_fixed_row_label_order_from_snapshot_config(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id

        output_path = TEST_ROOT / "exports" / "ordered-labels.xlsx"
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.review_session_service.export_session_revision(
            processing_run_id,
            session_revision=0,
            output_path=output_path,
        )

        worksheet = load_workbook(output_path)["Recap"]

        self.assertEqual(worksheet["A12"].value, "103 General FM")
        self.assertEqual(worksheet["A13"].value, "103 Foreman")
        self.assertEqual(worksheet["A14"].value, "103 Journeyman")

    def test_export_artifact_storage_uses_runtime_storage_seam_without_changing_workbook_behavior(self) -> None:
        processing_result = self._create_processing_run()
        artifact_store = LocalRuntimeFileStore(
            upload_root=TEST_ROOT / "runtime" / "uploads",
            export_root=TEST_ROOT / "runtime" / "exports",
        )
        service = ReviewSessionService(
            lineage_store=self.lineage_store,
            profile_execution_compatibility_adapter=self.profile_execution_compatibility_adapter,
            artifact_store=artifact_store,
            now_provider=lambda: self.created_at,
        )

        export_result = service.export_session_revision(
            processing_result.processing_run.processing_run_id,
            session_revision=0,
        )

        worksheet = load_workbook(export_result.output_path)["Recap"]
        resolved_payload = service.resolve_export_artifact_payload(export_result.export_artifact.export_artifact_id)

        self.assertTrue(export_result.export_artifact.storage_ref.startswith("exports/"))
        self.assertEqual(export_result.stored_artifact.storage_ref, export_result.export_artifact.storage_ref)
        self.assertEqual(export_result.stored_artifact.original_filename, "sample_report-recap-rev-0.xlsx")
        self.assertEqual(resolved_payload.file_path, export_result.output_path)
        self.assertEqual(worksheet["G27"].value, "Vendor A")

    def test_review_revision_advance_rolls_back_when_edit_insert_fails(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        review_session = self.review_session_service.open_review_session(processing_run_id).review_session
        self.lineage_store._connection.execute(
            """
            CREATE TRIGGER reviewed_record_edits_block_insert
            BEFORE INSERT ON reviewed_record_edits
            BEGIN
                SELECT RAISE(ABORT, 'block review edit insert');
            END;
            """
        )
        self.lineage_store._connection.commit()

        with self.assertRaisesRegex(Exception, "block review edit insert"):
            self.review_session_service.apply_review_edits(
                processing_run_id,
                [
                    PendingRecordEdit(
                        record_key="record-0",
                        changed_fields={"vendor_name_normalized": "Vendor B"},
                    )
                ],
            )

        persisted_session = self.lineage_store.get_review_session(review_session.review_session_id)
        persisted_edits = self.lineage_store.list_reviewed_record_edits(review_session.review_session_id)

        self.assertEqual(persisted_session.current_revision, 0)
        self.assertEqual(persisted_edits, [])

    def test_export_artifact_storage_is_cleaned_up_when_lineage_persistence_fails(self) -> None:
        processing_result = self._create_processing_run()
        artifact_store = TrackingArtifactStore()
        service = ReviewSessionService(
            lineage_store=self.lineage_store,
            profile_execution_compatibility_adapter=self.profile_execution_compatibility_adapter,
            artifact_store=artifact_store,
            now_provider=lambda: self.created_at,
        )

        with patch.object(
            self.lineage_store,
            "create_export_artifact",
            side_effect=RuntimeError("lineage insert failed"),
        ):
            with self.assertRaisesRegex(RuntimeError, "lineage insert failed"):
                service.export_session_revision(
                    processing_result.processing_run.processing_run_id,
                    session_revision=0,
                )

        self.assertEqual(len(artifact_store.saved_artifacts), 1)
        self.assertEqual(artifact_store.cleaned_storage_refs, [artifact_store.saved_artifacts[0].storage_ref])
        self.assertFalse(artifact_store.saved_artifacts[0].file_path.exists())

    def test_legacy_runs_are_marked_non_reproducible_and_fail_closed_for_historical_export(self) -> None:
        processing_result = self._create_processing_run()
        processing_run_id = processing_result.processing_run.processing_run_id
        self.lineage_store._connection.execute(
            """
            UPDATE profile_snapshots
            SET template_artifact_id = NULL,
                template_file_hash = NULL
            WHERE profile_snapshot_id = ?
            """,
            (processing_result.profile_snapshot.profile_snapshot_id,),
        )
        self.lineage_store._connection.commit()

        state = self.review_session_service.get_review_session_state(processing_run_id)

        self.assertEqual(state.historical_export_status.status_code, "legacy_non_reproducible")
        self.assertFalse(state.historical_export_status.is_reproducible)
        with self.assertRaises(HistoricalExportUnavailableError):
            self.review_session_service.export_session_revision(
                processing_run_id,
                session_revision=0,
                output_path=TEST_ROOT / "exports" / "legacy.xlsx",
            )

    def _create_processing_run(self):
        parsed_record = self._make_material_record(vendor_name_normalized="Vendor A")
        return self._create_processing_run_with_record(parsed_record)

    def _create_processing_run_with_record(self, parsed_record: Record):
        with patch(
            "services.review_workflow_service.parse_pdf",
            return_value=[parsed_record],
        ):
            return self.processing_run_service.create_processing_run(self.source_document_path)

    def _write_profile_bundle(self) -> None:
        profile_dir = TEST_ROOT / "profiles" / "default"
        self._write_json(
            profile_dir / "profile.json",
            {
                "profile_name": "default",
                "display_name": "Default Profile",
                "description": "Default test profile",
                "version": "1.0",
                "template_filename": "recap_template.xlsx",
                "is_active": False,
            },
        )
        self._write_json(profile_dir / "labor_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "equipment_mapping.json", {"raw_mappings": {}, "saved_mappings": []})
        self._write_json(profile_dir / "phase_mapping.json", {"50": "MATERIAL"})
        self._write_json(profile_dir / "vendor_normalization.json", {})
        self._write_json(
            profile_dir / "input_model.json",
            {"report_type": "vista_job_cost", "section_headers": {}},
        )
        self._write_json(
            profile_dir / "target_labor_classifications.json",
            {
                "slots": [
                    {"slot_id": "labor_1", "label": "103 General FM", "active": True},
                    {"slot_id": "labor_2", "label": "103 Foreman", "active": True},
                    {"slot_id": "labor_3", "label": "103 Journeyman", "active": True},
                ],
                "classifications": ["103 General FM", "103 Foreman", "103 Journeyman"],
            },
        )
        self._write_json(
            profile_dir / "target_equipment_classifications.json",
            {
                "slots": [
                    {"slot_id": "equipment_1", "label": "Pick-up Truck", "active": True},
                ],
                "classifications": ["Pick-up Truck"],
            },
        )
        self._write_json(
            profile_dir / "rates.json",
            {"labor_rates": {}, "equipment_rates": {}},
        )
        self._write_json(
            profile_dir / "review_rules.json",
            {"default_omit_rules": []},
        )
        self._write_json(
            profile_dir / "recap_template_map.json",
            {
                "worksheet_name": "Recap",
                "header_fields": {
                    "project": {"cell": "B6"},
                    "job_number": {"cell": "H6"},
                },
                "labor_rows": {
                    "103 General FM": {
                        "st_hours": "B12",
                        "ot_hours": "C12",
                        "dt_hours": "D12",
                        "st_rate": "E12",
                        "ot_rate": "F12",
                        "dt_rate": "G12",
                    },
                    "103 Foreman": {
                        "st_hours": "B13",
                        "ot_hours": "C13",
                        "dt_hours": "D13",
                        "st_rate": "E13",
                        "ot_rate": "F13",
                        "dt_rate": "G13",
                    },
                    "103 Journeyman": {
                        "st_hours": "B14",
                        "ot_hours": "C14",
                        "dt_hours": "D14",
                        "st_rate": "E14",
                        "ot_rate": "F14",
                        "dt_rate": "G14",
                    }
                },
                "equipment_rows": {
                    "Pick-up Truck": {"hours_qty": "B32", "rate": "D32"}
                },
                "materials_section": {
                    "start_row": 27,
                    "end_row": 41,
                    "columns": {"name": "G", "amount": "H"},
                },
                "subcontractors_section": {
                    "start_row": 46,
                    "end_row": 50,
                    "columns": {"name": "A", "amount": "C"},
                },
                "permits_fees_section": {
                    "start_row": 55,
                    "end_row": 56,
                    "columns": {"description": "A", "amount": "C"},
                },
                "police_detail_section": {
                    "start_row": 61,
                    "end_row": 62,
                    "columns": {"description": "A", "amount": "C"},
                },
                "sales_tax_area": {
                    "rate_label_cell": "G60",
                    "rate_input_cell": "H60",
                    "amount_label_cell": "G61",
                    "amount_formula_cell": "H61",
                    "material_total_cell": "H54",
                },
            },
        )
        self._create_template(profile_dir / "recap_template.xlsx")

    def _create_template(self, path: Path, *, marker_value: str = "Template V1") -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Recap"

        for cell, value in {
            "A1": marker_value,
            "A6": "Project",
            "G6": "Job Number",
            "H23": "=SUM(H12:H22)",
            "A25": "EQUIPMENT",
            "A26": "Category",
            "B26": "Hours / Qty",
            "D26": "Rate",
            "G25": "MATERIALS",
            "G26": "Vendor",
            "H26": "Amount",
            "E42": "=SUM(E27:E41)",
            "H42": "=SUM(H27:H41)",
            "C51": "=SUM(C46:C50)",
            "C57": "=SUM(C55:C56)",
            "C63": "=SUM(C61:C62)",
            "F58": 0,
        }.items():
            worksheet[cell] = value

        workbook.save(path)

    def _make_material_record(self, *, vendor_name_normalized: str) -> Record:
        return Record(
            record_type=MATERIAL,
            phase_code="50",
            raw_description="Material line",
            cost=100.0,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name="Vendor A",
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Material source",
            record_type_normalized=MATERIAL,
            recap_labor_classification=None,
            vendor_name_normalized=vendor_name_normalized,
        )

    def _make_missing_vendor_material_record(self) -> Record:
        return Record(
            record_type=MATERIAL,
            phase_code="50",
            raw_description="Material line",
            cost=100.0,
            hours=None,
            hour_type=None,
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=[],
            job_number="JOB-100",
            job_name="Sample Project",
            source_page=1,
            source_line_text="Material source",
            record_type_normalized=MATERIAL,
            recap_labor_classification=None,
            vendor_name_normalized=None,
        )

    def _make_unmapped_equipment_record(self, *, warnings: list[str] | None = None) -> Record:
        return Record(
            record_type=EQUIPMENT,
            phase_code=None,
            raw_description="627/2025 crane truck",
            cost=100.0,
            hours=8.0,
            hour_type="EA",
            union_code=None,
            labor_class_raw=None,
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description="627/2025 crane truck",
            equipment_category=None,
            confidence=0.9,
            warnings=list(warnings or []),
            source_page=1,
            source_line_text="Equipment source",
            record_type_normalized=None,
        )

    def _make_unmapped_labor_record(self, *, warnings: list[str] | None = None) -> Record:
        return Record(
            record_type=LABOR,
            phase_code="20",
            raw_description="Labor line",
            cost=100.0,
            hours=8.0,
            hour_type="ST",
            union_code="103",
            labor_class_raw="ZZ",
            labor_class_normalized=None,
            vendor_name=None,
            equipment_description=None,
            equipment_category=None,
            confidence=0.9,
            warnings=list(warnings or []),
            source_page=1,
            source_line_text="Labor source",
            record_type_normalized=None,
            recap_labor_classification=None,
        )

    def _write_json(self, path: Path, payload: object) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


class TrackingArtifactStore:
    """Minimal runtime storage double that records export cleanup activity."""

    def __init__(self) -> None:
        self._tmpdir = tempfile.TemporaryDirectory(prefix="review-session-artifacts-")
        self.saved_artifacts: list[StoredArtifact] = []
        self.cleaned_storage_refs: list[str] = []

    def save_export_artifact(
        self,
        *,
        processing_run_id: str,
        session_revision: int,
        original_filename: str,
        content_bytes: bytes,
        content_type: str | None = None,
    ) -> StoredArtifact:
        artifact_dir = Path(self._tmpdir.name) / f"{processing_run_id.replace(':', '-')}-{session_revision}-{len(self.saved_artifacts)}"
        artifact_dir.mkdir(parents=True, exist_ok=False)
        artifact_path = artifact_dir / original_filename
        artifact_path.write_bytes(content_bytes)
        stored_artifact = StoredArtifact(
            storage_ref=f"exports/{artifact_dir.name}/{original_filename}",
            original_filename=original_filename,
            content_type=content_type or "application/octet-stream",
            file_size_bytes=len(content_bytes),
            file_path=artifact_path,
        )
        self.saved_artifacts.append(stored_artifact)
        return stored_artifact

    def get_export_artifact(self, storage_ref: str) -> StoredArtifact:
        for stored_artifact in self.saved_artifacts:
            if stored_artifact.storage_ref == storage_ref:
                return stored_artifact
        raise FileNotFoundError(storage_ref)

    def delete_export_artifact(self, storage_ref: str) -> None:
        self.cleaned_storage_refs.append(storage_ref)
        for stored_artifact in self.saved_artifacts:
            if stored_artifact.storage_ref == storage_ref:
                shutil.rmtree(stored_artifact.file_path.parent, ignore_errors=True)
                return


if __name__ == "__main__":
    unittest.main()
